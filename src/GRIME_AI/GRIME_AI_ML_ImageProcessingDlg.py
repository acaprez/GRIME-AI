#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# Author: John Edward Stranzl, Jr.
# Affiliation(s): University of Nebraska-Lincoln, Blade Vision Systems, LLC
# Contact: jstranzl2@huskers.unl.edu, johnstranzl@gmail.com
# Created: Mar 6, 2022
# License: Apache License, Version 2.0, http://www.apache.org/licenses/LICENSE-2.0

import os
import getpass
import json
import csv
import cv2
from pathlib import Path
from promptlib import Files
from typing import List, Tuple, Dict

from PyQt5.QtGui import QPixmap, QIcon, QPainter, QColor, QImage
from PyQt5.QtCore import Qt, pyqtSignal, QTimer
from PyQt5.QtWidgets import QDialog, QFileDialog, QListWidgetItem, QAbstractItemView, QSizePolicy, QListWidget, QMessageBox, QDialogButtonBox
from PyQt5.uic import loadUi
from PyQt5 import QtCore, QtWidgets

import matplotlib
matplotlib.use("Qt5Agg")      # <<< FORCE Qt5Agg backend for PyQt5

from GRIME_AI.GRIME_AI_Save_Utils import GRIME_AI_Save_Utils
from GRIME_AI.GRIME_AI_Save_Utils import JsonEditor
from GRIME_AI.coco_generator import CocoGenerator
from GRIME_AI.GRIME_AI_ImageAnnotatorDlg import ImageAnnotatorDialog
from GRIME_AI.GRIME_AI_CSS_Styles import BUTTON_CSS_STEEL_BLUE, BUTTON_CSS_DARK_RED

import torch


# ======================================================================================================================
# ======================================================================================================================
#  =====     =====     =====     =====     =====   MODULE LEVEL HELPERS    =====     =====     =====     =====     =====
# ======================================================================================================================
# ======================================================================================================================
def _check_folder(folder: Path) -> Tuple[bool, List[str]]:
    """
    Validates a single folder by:
      1. Finding at least one .json COCO file and some .jpg/.jpeg images.
      2. Parsing the JSON’s "images" list of dicts to pull out file_name.
      3. Verifying every listed file_name exists in that folder.

    Returns (is_valid, missing_files_list).
    """
    # 1) List JSONs and JPGs via os.scandir
    jsons = [e.name for e in os.scandir(folder)
             if e.is_file() and e.name.lower().endswith(".json")]
    jpgs  = {e.name for e in os.scandir(folder)
             if e.is_file() and e.name.lower().endswith((".jpg", ".jpeg"))}

    print(f"Scanning `{folder}` -> JSONs: {jsons}, JPGs: {list(jpgs)[:5]}…")  # debug

    if not jsons or not jpgs:
        return False, []

    # 2) Load the first JSON file
    path_json = folder / jsons[0]
    try:
        data = json.loads(path_json.read_text(encoding="utf-8"))
    except Exception as e:
        return False, [f"Cannot parse {jsons[0]}: {e}"]

    # 3) Extract expected filenames from COCO "images" list
    raw_images = data.get("images")
    if not isinstance(raw_images, list):
        return False, [f"'images' key missing or not a list in {jsons[0]}"]

    expected_files = []
    for item in raw_images:
        if isinstance(item, dict):
            fname = item.get("file_name") or item.get("filename")
            if not fname:
                return False, [f"Missing 'file_name' in entry: {item}"]
            expected_files.append(Path(fname).name)
        elif isinstance(item, str):
            expected_files.append(item)
        else:
            return False, [f"Unsupported image entry type: {type(item)}"]

    # 4) Compare against the actual JPGs on disk
    missing = [f for f in expected_files if f not in jpgs]
    if missing:
        return False, missing

    return True, []


def _iter_dirs(root: Path):
    """
    Recursively yield every subdirectory under root using os.scandir.
    """
    for entry in os.scandir(root):
        if entry.is_dir():
            sub = Path(entry.path)
            yield sub
            yield from _iter_dirs(sub)


# ======================================================================================================================
# ======================================================================================================================
#  =====     =====     =====     =====     =====     =====     =====     =====     =====     =====     =====     =====
# ======================================================================================================================
# ======================================================================================================================
# Custom ListWidget classes with drag and drop support.
class DraggableListWidget(QListWidget):
    def mimeData(self, items):
        mimeData = QtCore.QMimeData()
        texts = "\n".join(sorted(set(item.text() for item in self.selectedItems())))
        mimeData.setText(texts)
        return mimeData


# ======================================================================================================================
# ======================================================================================================================
#  =====     =====     =====     =====     =====     =====     =====     =====     =====     =====     =====     =====
# ======================================================================================================================
# ======================================================================================================================
class DroppableListWidget(QListWidget):
    def dropEvent(self, event):
        if event.mimeData().hasText():
            text = event.mimeData().text()
            # In case multiple items were dragged:
            items_to_drop = [line.strip() for line in text.splitlines() if line.strip()]
            dlg = self.parent()
            if dlg is not None:
                for item_text in items_to_drop:
                    # Remove matching item from available list.
                    available = dlg.listWidget_availableFolders
                    for idx in range(available.count()):
                        avail_item = available.item(idx)
                        if avail_item.text() == item_text:
                            available.takeItem(idx)
                            break
                    # Add the item if not already transferred.
                    if item_text not in dlg.transferred_items:
                        self.addItem(item_text)
                        dlg.transferred_items.add(item_text)
                        print(f"Dragged '{item_text}' from available to selected folders via drop.")
                dlg.updateTrainButtonState()  # Update Train button after drop.
            event.accept()
        else:
            event.ignore()


# ======================================================================================================================
# ======================================================================================================================
#  =====     =====     =====     =====     class GRIME_AI_ML_ImageProcessingDlg      =====     =====     =====     =====
# ======================================================================================================================
# ======================================================================================================================
class GRIME_AI_ML_ImageProcessingDlg(QDialog):

    ml_train_signal = pyqtSignal()
    ml_segment_signal = pyqtSignal()

    def __init__(self, parent=None):
        super().__init__(parent)

        UI_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)),'ui',"QDialog_ML_ImageProcessing.ui")
        loadUi(UI_FILE, self)

        self._pendingThumbnails = []
        self._batchSize = 10  # number of thumbs per batch
        self._batchDelay = 50  # ms between batches
        self._loadToken = 0  # to cancel stale batches

        # --------------------------------------------------------------------------------------------------------------
        ###JES hide GRIME AI annotation/labeling; users must use CVAT
        # JES - The Image Annotation/Labeling functionality is currently in development and not intended for public release.
        # JES - Access is restricted to the development team. While it may be technically possible to circumvent
        # JES - these restrictions, GRIME Lab and its developers accept no liability or responsibility for any
        # JES - consequences arising from such actions.
        #
        # JES - Licensed under the Apache License, Version 2.0 (the "License");
        # JES - you may not use this file except in compliance with the License.
        # JES - You may obtain a copy of the License at:
        # JES -     http://www.apache.org/licenses/LICENSE-2.0
        # --------------------------------------------------------------------------------------------------------------
        tb = self.tabWidget.tabBar()
        if getpass.getuser() == "johns" or getpass.getuser() == "tgilmore10":
            tb.setTabVisible(3, True)
        else:
            tb.setTabVisible(3, False)


        # Initialize per-image annotation store
        # Keys: full image path, Values: list of {type, points}
        self.annotation_store: Dict[str, List[Dict]] = {}

        # --- Filmstrip one‐row configuration ---
        filmstrip = self.listWidget_annotationFilmstrip

        # 1. Layout: Left‐to‐right, no wrapping, static movement
        filmstrip.setFlow(QtWidgets.QListView.LeftToRight)
        filmstrip.setWrapping(False)
        filmstrip.setResizeMode(QtWidgets.QListView.Adjust)
        filmstrip.setMovement(QtWidgets.QListView.Static)

        # 2. Scroll bars: only horizontal
        filmstrip.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        filmstrip.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarPolicy.ScrollBarAlwaysOn)

        # 3. Fix height to exactly one icon row
        icon_h = filmstrip.iconSize().height()
        # account for widget frame borders
        frame = filmstrip.frameWidth() * 2
        # if you want a tiny gap above/below, you can include spacing()
        spacing = filmstrip.spacing()
        filmstrip.setFixedHeight(icon_h + frame + spacing)

        # Optional: Prevent the layout from stretching it vertically
        filmstrip.setSizePolicy(
            QtWidgets.QSizePolicy.Expanding,
            QtWidgets.QSizePolicy.Fixed
        )

        # annotation tab state
        self._annotation_image_paths = []

        le = self.lineEdit_annotationCategory
        lst = self.listWidget_annotationCategories

        # Enter → add to list
        le.returnPressed.connect(lambda: self._add_category(le, lst))

        # ---
        # --- Segment Images Tab Layout
        # ---
        layout = self.horizontalLayoutSegmentImages
        layout.setStretch(0, 4)  # left content area
        layout.setStretch(1, 1)  # right 'Labels' group box

        # filmstrip widget
        lw = self.listWidget_filmstrip

        # 1) no wrapping, no internal spacing
        lw.setWrapping(False)
        lw.setSpacing(0)

        # 2) remove any margins between viewport and frame
        lw.setContentsMargins(0, 0, 0, 0)
        lw.setViewportMargins(0, 0, 0, 0)

        # 3) remove the frame entirely
        from PyQt5.QtWidgets import QFrame

        lw.setFrameShape(QFrame.NoFrame)

        # 4) fix the height to exactly the icon height
        icon_h = lw.iconSize().height()
        lw.setFixedHeight(icon_h)

        self.lineEdit_siteName.editingFinished.connect(self.save_site_name_to_json)

        # Initialize tracking variables.
        self.transferred_items = set()
        self.original_folders = []
        self.selected_label_categories = []
        self.categories_available = False

        # Call helper methods.
        #self.setup_from_config_file()
        self.setup_custom_list_widgets()
        self.setup_ui_properties()
        self.setup_connections()
        self.setup_drag_and_drop()
        self.populate_segment_images_tab()
        self.updateTrainButtonState()

        model_path = self.lineEdit_segmentation_model_file.text().strip()
        self.updateSegmentButtonState(model_path)

        # install event filter on the annotation‐image label
        self.label_imageAnnotation.installEventFilter(self)

        if self.lineEdit_model_training_images_path.text():
            self.populate_available_folders()

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def save_site_name_to_json(self):
        import os
        import json

        settings_folder = GRIME_AI_Save_Utils().get_settings_folder()
        CONFIG_FILENAME = "site_config.json"
        config_file = os.path.join(settings_folder, CONFIG_FILENAME)

        # Load current settings if file exists
        if os.path.exists(config_file):
            with open(config_file, "r", encoding="utf-8") as f:
                settings = json.load(f)
        else:
            settings = {}

        # Update with the current SiteName value
        settings["siteName"] = self.lineEdit_siteName.text()

        # Write back to the config file
        with open(config_file, "w", encoding="utf-8") as f:
            json.dump(settings, f, indent=4)

        print(f"Updated siteName in {config_file} to '{settings['siteName']}'")

    # ------------------------------------------------------------------------------------------------------------------
    # Debug Helper: Dump all stored annotations
    # ------------------------------------------------------------------------------------------------------------------
    def print_all_annotations(self):
        """
        Prints every image path and the count/details of shapes stored
        in self.annotation_store for debugging.
        """
        if not self.annotation_store:
            print("No annotations have been recorded yet.")
            return


        for img_path, shapes in self.annotation_store.items():
            print(f"\nAnnotations for {img_path}:")
            for idx, shape in enumerate(shapes, start=1):
                pts = shape['points']
                print(f"  {idx}. {shape['type']} with {len(pts)} points → {pts}")


    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def load_labels_from_annotation(self, folder_path):
        annotation_file = os.path.join(folder_path, "instances_default.json")
        if not os.path.exists(annotation_file):
            return []

        with open(annotation_file, "r", encoding="utf-8") as f:
            data = json.load(f)

        labels = set()
        if "annotations" in data and "categories" in data:
            for cat in data["categories"]:
                labels.add(f"{cat['id']} - {cat['name']}")
        return sorted(labels)

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def update_annotation_listbox(self, listbox_widget, folder_path):
        labels = self.load_labels_from_annotation(folder_path)
        listbox_widget.clear()
        listbox_widget.addItems(labels)

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def _add_category(self, lineedit, listwidget):
        text = lineedit.text().strip()
        if text and not listwidget.findItems(text, Qt.MatchExactly):
            listwidget.addItem(text)
        lineedit.clear()

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def _openAnnotator(self):
        pm = self.label_imageAnnotation.pixmap()
        if pm is None:
            QMessageBox.warning(self, "No Image", "No image is loaded.")
            return

        # Determine drawing mode
        if self.radioButton_boundingBox.isChecked():
            mode = 'bbox'
        elif self.radioButton_polylineClick.isChecked():
            mode = 'click'
        else:
            mode = 'drag'

        # Figure out which category is selected in your Annotation tab
        current = self.listWidget_annotationCategories.currentItem()
        if current:
            name = current.text()
            idx  = self.listWidget_annotationCategories.currentRow() + 1
        else:
            name = "<unknown>"
            idx  = -1

        # Load any existing shapes from disk
        current_item = self.listWidget_annotationFilmstrip.currentItem()
        img_idx = current_item.data(Qt.UserRole)
        img_path = self._annotation_image_paths[img_idx]
        img_path = os.path.normpath(img_path)  # normalizes slashes
        img_path = img_path.replace("\\", "/")  # force-forward slashes

        # DEBUG: Print resolved image path and store state
        print(f"[DEBUG] Current image index: {img_idx}")
        print(f"[DEBUG] Resolved image path:\n  {img_path}")
        print(f"[DEBUG] Annotation store keys:\n  {list(self.annotation_store.keys())}")

        ###JES self.load_annotations_for_image(img_path)

        # Launch dialog, passing along the label ID & name
        dlg = ImageAnnotatorDialog(
            pm,
            mode=mode,
            label={"id": idx, "name": name},
            parent=self
        )

        # Preload into the annotator
        existing = self.annotation_store.get(img_path)
        print(f"[DEBUG] Loaded {len(existing) if existing else 0} shapes for this image.")
        print(f"[DEBUG] Opening image: {img_path}")
        print(f"[DEBUG] Stored shapes: {len(existing) if existing else 0}")
        if existing:
            dlg.setAnnotations(existing)

        if dlg.exec_():
            shapes = dlg.getAnnotations()
            self.annotation_store[img_path] = shapes
            self.save_annotations_for_image(img_path)
            print(f"[DEBUG] Saved {len(shapes)} shapes for {img_path}")

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def save_annotations_for_image(self, img_path: str):
        """
        Writes the annotation JSON for a single image into
        an 'annotations/' subfolder next to the image.
        """
        shapes = self.annotation_store.get(img_path, [])
        if not shapes:
            return

        img_p = Path(img_path)
        ann_folder = img_p.parent / "annotations"
        try:
            ann_folder.mkdir(parents=True, exist_ok=True)
        except Exception as e:
            print(f"Could not create annotation folder {ann_folder}: {e}")

        ann_file = ann_folder / f"{img_p.stem}.anno.json"
        try:
            with open(ann_file, "w", encoding="utf-8") as f:
                json.dump(shapes, f, indent=2)
            print(f"Saved annotations to {ann_file}")
        except Exception as e:
            print(f"Failed to save annotations to {ann_file}: {e}")

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def load_annotations_for_image(self, img_path: str):
        """
        Loads the annotation JSON for a single image from
        the 'annotations/' subfolder, if it exists.
        """
        img_p = Path(img_path)
        ann_file = img_p.parent / "annotations" / f"{img_p.stem}.anno.json"
        if not ann_file.exists():
            return

        try:
            with open(ann_file, "r", encoding="utf-8") as f:
                shapes = json.load(f)
            # store under the image key (string)
            self.annotation_store[img_path] = shapes
            print(f"Loaded annotations from {ann_file}")
        except Exception as e:
            print(f"Failed to load annotations from {ann_file}: {e}")

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def _refresh_annotation_categories(self):
        """
        Scan self.annotation_store for all loaded shapes,
        extract unique (id, name) pairs and populate the category list.
        """
        unique = {}
        for shapes in self.annotation_store.values():
            for s in shapes:
                lab = s.get("label", {})
                cid = lab.get("id", None)
                name = lab.get("name", None)
                if cid is not None and name:
                    unique[cid] = name

        self.listWidget_annotationCategories.clear()
        for cid, name in sorted(unique.items()):
            self.listWidget_annotationCategories.addItem(name)

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def save_coco_for_all_images(self, base_image_folder: str):
        """
        Walks through the annotation_store for every image,
        builds a single COCO 1.0 JSON, and writes it to
        'annotations/coco_all.json' under base_image_folder.
        """
        # Prepare output folder
        base_p = Path(base_image_folder)
        ann_folder = base_p / "annotations"
        ann_folder.mkdir(parents=True, exist_ok=True)

        coco = {
            "info": {
                "description": "Combined dataset",
                "version": "1.0",
                "year": 2025
            },
            "licenses": [],
            "images": [],
            "annotations": [],
            "categories": []
        }

        image_id_map = {}  # map image path -> image_id
        category_map = {}  # map cat_id -> cat_name
        ann_id = 1

        # Helper to compute polygon area via shoelace
        def polygon_area(coords):
            area = 0.0
            n = len(coords)
            for i in range(n):
                x1, y1 = coords[i]
                x2, y2 = coords[(i + 1) % n]
                area += x1 * y2 - x2 * y1
            return abs(area) * 0.5

        # 1) Iterate through every image in your store
        for idx, (img_path, shapes) in enumerate(self.annotation_store.items(), start=1):
            if not shapes:
                continue

            img_p = Path(img_path)
            # assign a unique image_id
            image_id = idx
            image_id_map[img_path] = image_id

            # read size
            img = cv2.imread(str(img_p))
            if img is None:
                print(f"Warning: cannot load {img_p}, skipping.")
                continue
            h, w = img.shape[:2]

            # append image entry
            coco["images"].append({
                "id": image_id,
                "file_name": img_p.name,
                "width": w,
                "height": h
            })

            # collect categories seen in this image
            for shape in shapes:
                cid = shape["label"]["id"]
                cname = shape["label"]["name"]
                category_map[cid] = cname

            # build annotation entries
            for shape in shapes:
                pts = shape["points"]
                # segmentation: flatten [x1,y1,x2,y2,...]
                seg = []
                for p in pts:
                    seg.extend([p.x(), p.y()])

                xs = [p.x() for p in pts]
                ys = [p.y() for p in pts]
                x_min, x_max = min(xs), max(xs)
                y_min, y_max = min(ys), max(ys)
                bbox = [x_min, y_min, x_max - x_min, y_max - y_min]
                area = polygon_area([(p.x(), p.y()) for p in pts])

                coco["annotations"].append({
                    "id": ann_id,
                    "image_id": image_id,
                    "category_id": shape["label"]["id"],
                    "segmentation": [seg],
                    "bbox": bbox,
                    "area": area,
                    "iscrowd": 0
                })
                ann_id += 1

        # 2) Finalize category list
        coco["categories"] = [
            {"id": cid, "name": cname, "supercategory": ""}
            for cid, cname in category_map.items()
        ]

        # 3) Write out the combined COCO JSON
        out_file = ann_folder / "coco_all.json"
        try:
            with open(out_file, "w", encoding="utf-8") as f:
                json.dump(coco, f, indent=2)
            print(f"Saved global COCO file to {out_file}")
        except Exception as e:
            print(f"Failed to write global COCO JSON: {e}")

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def setup_from_config_file(self):
        """
        Initialize dialog controls from a configuration dictionary.
        """
        self.lineEdit_siteName.setText(self.config.get("siteName", ""))
        learningRates = self.config.get("learningRates", [])
        lr_str = ", ".join(str(x) for x in learningRates)
        self.lineEdit_learningRates.setText(lr_str)

        optimizer = self.config.get("Optimizer", "")
        idx = self.comboBox_optimizer.findText(optimizer)
        if idx >= 0:
            self.comboBox_optimizer.setCurrentIndex(idx)

        loss_function = self.config.get("loss_function", "")
        idx = self.comboBox_lossFunction.findText(loss_function)
        if idx >= 0:
            self.comboBox_lossFunction.setCurrentIndex(idx)

        self.doubleSpinBox_weightDecay.setValue(self.config.get("weight_decay", 0.0))
        self.spinBox_epochs.setValue(self.config.get("number_of_epochs", 0))
        self.spinBox_batchSize.setValue(self.config.get("batch_size", 0))
        self.spinBox_saveFrequency.setValue(self.config.get("save_model_frequency", 0))
        self.spinBox_validationFrequency.setValue(self.config.get("validation_frequency", 0))
        self.checkBox_earlyStopping.setChecked(self.config.get("early_stopping", False))
        self.spinBox_patience.setValue(self.config.get("patience", 0))

        device = self.config.get("device", "")
        idx = self.comboBox_device.findText(device)
        if idx >= 0:
            self.comboBox_device.setCurrentIndex(idx)

        self.lineEdit_segmentation_model_file.setText(self.config.get("segmentation_model_file", ""))
        self.lineEdit_segmentation_images_folder.setText(self.config.get("segmentation_images_folder", ""))
        self.checkBox_saveModelMasks.setChecked(self.config.get("save_model_masks", True))
        self.checkBox_copyOriginalModelImage.setChecked(self.config.get("copy_original_model_image", True))

        load_model_conf = self.config.get("load_model", {})
        if load_model_conf:
            model_path = load_model_conf.get("MODEL", "")
            if model_path:
                self.lineEdit_segmentation_model_file.setText(model_path)
                print("Populated segmentation model file from JSON (MODEL):", model_path)
            input_dir = load_model_conf.get("INPUT_DIR", "")
            if input_dir:
                self.lineEdit_segmentation_images_folder.setText(input_dir)
                print("Populated segmentation images folder from JSON (INPUT_DIR):", input_dir)

        self.current_path = self.config.get("Path", None)

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def setup_custom_list_widgets(self):
        """Replace default list widgets with custom draggable/droppable ones."""
        self.listWidget_availableFolders.__class__ = DraggableListWidget
        self.listWidget_selectedFolders.__class__ = DroppableListWidget

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def setup_ui_properties(self):
        """Set size policies and layout stretch factors."""
        self.tabWidget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.listWidget_availableFolders.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.listWidget_selectedFolders.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.listWidget_availableFolders.setMinimumHeight(200)
        self.listWidget_selectedFolders.setMinimumHeight(200)
        self.adjustSize()
        self.setMinimumSize(self.size())
        self.verticalTabParametersLayout.setStretch(0, 1)
        self.verticalTabParametersLayout.setStretch(1, 0)
        self.horizontalMainLayout.setStretch(0, 1)
        self.horizontalMainLayout.setStretch(1, 3)
        self.horizontalListLayout.setStretch(0, 1)
        self.horizontalListLayout.setStretch(1, 0)
        self.horizontalListLayout.setStretch(2, 1)

        # Set stylesheet for the tabs to change color when a tab is selected.
        self.tabWidget.setStyleSheet("""
            QTabBar::tab {
                background-color: white;
                color: black;
            }
            QTabBar::tab:selected {
                background-color: steelblue;
                color: white;
            }
        """)

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def setup_connections(self):
        """Connect signals with their slot methods."""
        # Folder management signals.
        self.pushButton_browse_model_training_images_folder.clicked.connect(self.browse_model_training_images_folder)
        model_training_image_folder = JsonEditor().getValue("Model_Training_Images_Folder")
        if model_training_image_folder:
            self.lineEdit_model_training_images_path.setText(model_training_image_folder)
        self.lineEdit_model_training_images_path.editingFinished.connect(self.populate_available_folders)

        self.pushButton_moveRight.clicked.connect(self.move_to_right)
        self.pushButton_moveRight.setStyleSheet(BUTTON_CSS_STEEL_BLUE)

        self.pushButton_moveLeft.clicked.connect(self.move_to_left)
        self.pushButton_moveLeft.setStyleSheet(BUTTON_CSS_STEEL_BLUE)

        self.pushButton_reset.clicked.connect(self.reset_lists)
        self.pushButton_reset.setStyleSheet(BUTTON_CSS_DARK_RED)

        self.pushButton_train.clicked.connect(self.train)
        self.pushButton_train.setStyleSheet(BUTTON_CSS_STEEL_BLUE)

        self.listWidget_availableFolders.itemDoubleClicked.connect(self.handle_left_item_doubleclick)
        self.listWidget_selectedFolders.itemDoubleClicked.connect(self.handle_right_item_doubleclick)

        # Data Annotation tab connections
        self.pushButton_annotationBrowse.clicked.connect(self.browse_annotation_folder)
        self.pushButton_annotationBrowse.setStyleSheet(BUTTON_CSS_STEEL_BLUE)

        self.listWidget_annotationFilmstrip.itemClicked.connect(self.display_annotation_image)

        # Segment Images tab signals.
        self.pushButton_Select_Model.clicked.connect(self.select_segmentation_model)
        self.pushButton_Select_Model.setStyleSheet(BUTTON_CSS_STEEL_BLUE)

        self.pushButton_Select_Images_Folder.clicked.connect(self.select_segmentation_images_folder)
        self.pushButton_Select_Images_Folder.setStyleSheet(BUTTON_CSS_STEEL_BLUE)

        self.pushButton_Segment.clicked.connect(self.segment_images)
        self.pushButton_Segment.setStyleSheet(BUTTON_CSS_STEEL_BLUE)

        self.lineEdit_segmentation_model_file.textChanged.connect(self.updateSegmentButtonState)
        self.lineEdit_segmentation_images_folder.textChanged.connect(self.updateSegmentButtonState)
        self.checkBox_saveModelMasks.toggled.connect(self.on_save_masks_toggled)
        self.checkBox_copyOriginalModelImage.toggled.connect(self.on_copy_original_toggled)

        # ### COCO GENERATION TAB CONNECTIONS ###
        # Connect COCO tab widgets (ensure these names match the UI file).
        self.lineEdit_cocoFolder.textChanged.connect(self.updateCOCOButtonState)
        self.pushButton_cocoBrowse.clicked.connect(self.selectCocoFolder)
        self.pushButton_cocoBrowse.setStyleSheet(BUTTON_CSS_STEEL_BLUE)

        self.lineEdit_maskFile.textChanged.connect(self.updateCOCOButtonState)
        self.pushButton_maskBrowse.clicked.connect(self.selectMaskFile)
        self.pushButton_maskBrowse.setStyleSheet(BUTTON_CSS_STEEL_BLUE)

        self.checkBox_singleMask.toggled.connect(self.updateMaskFieldState)
        self.pushButton_generateCOCO.clicked.connect(self.generateCOCOAnnotations)
        self.pushButton_generateCOCO.setStyleSheet(BUTTON_CSS_STEEL_BLUE)

        self.updateMaskFieldState(self.checkBox_singleMask.isChecked())
        self.updateCOCOButtonState()
        # ### END NEW COCO CONNECTIONS ###

        # <<<< Connection for ROI Analyzer Analyze button >>>>
        self.lineEdit_ROI_images_folder.editingFinished.connect(self._on_roi_images_folder_changed)
        self.pushButton_browse_ROI_images_folder.clicked.connect(self.browse_ROI_images_folder)
        self.pushButton_browse_ROI_images_folder.setStyleSheet(BUTTON_CSS_STEEL_BLUE)

        self.pushButton_analyze.clicked.connect(self.analyze_roi)
        self.pushButton_analyze.setStyleSheet(BUTTON_CSS_STEEL_BLUE)

        self.pushButton_extract_ROI_features.clicked.connect(self.extract_ROI_features)
        self.pushButton_extract_ROI_features.setStyleSheet(BUTTON_CSS_STEEL_BLUE)

        self.listWidget_filmstrip.itemClicked.connect(self.on_filmstrip_item_clicked)

        self.num_clusters = self.spinBox_numClusters.value()
        self.spinBox_numClusters.valueChanged.connect(self._on_num_clusters_changed)

        self.buttonBox_close.rejected.connect(self.reject)

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def reject(self):
        # DO NOTHING. LET IT CLOSE. IF THE CALLING PROGRAM CREATED THE DIALOG USING EXEC, THE CALLING INSTANTIATING
        # PROGRTAM CAN INSPECT THE RETURN RESULT
        super().reject()

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def _on_num_clusters_changed(self, value):
        self.num_clusters = value

        current = self.listWidget_filmstrip.currentItem()
        if current:
            # re-run analysis on the highlighted image
            self.on_filmstrip_item_clicked(current)
        else:
            # fallback: rerun the initial analysis sequence
            self.analyze_roi()

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def on_save_masks_toggled(self, checked: bool):
        print(f"Save Masks checkbox toggled: {checked}")
        # Add any logic here to update internal state or UI

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def on_copy_original_toggled(self, checked: bool):
        print(f"Copy Original Image checkbox toggled: {checked}")
        # Add any logic here to update internal state or UI

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def on_filmstrip_item_clicked(self, item: QListWidgetItem):
        """
        When a thumbnail is clicked, re-run ROI analysis on that image/mask pair
        and update label_displayImages with the composite plot + top-3 swatches.
        """
        # Retrieve the index we stored in populate_filmstrip
        idx = item.data(Qt.UserRole)
        orig_path, mask_path = self._pairs[idx]

        # grab the user-selected cluster count
        n_clusters = self.spinBox_numClusters.value()

        # Run analysis for this specific pair
        from GRIME_AI.GRIME_AI_ROI_Analyzer import GRIME_AI_ROI_Analyzer
        analyzer = GRIME_AI_ROI_Analyzer(orig_path, mask_path, clusters=n_clusters)
        analyzer.run_analysis()

        # ─── populate metric fields ───
        self.lineEdit_intensity.setText(f"{analyzer.roi_intensity:.2f}")
        self.lineEdit_entropy.setText(f"{analyzer.roi_entropy:.4f}")
        self.lineEdit_Texture.setText(f"{analyzer.roi_texture:.4f}")
        self.lineEdit_GLI.setText(f"{analyzer.mean_gli:.2f}")
        self.lineEdit_GCC.setText(f"{analyzer.mean_gcc:.2f}")

        # Display composite+metrics plot
        composite_pix = analyzer.get_results_pixmap()
        self.label_displayImages.setPixmap(composite_pix)

        # Overlay top-3 dominant color swatches
        swatches = analyzer.dominant_rgb_list[:3]
        self._draw_color_swatches_on_label(swatches, swatch_size=100)

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def populate_filmstrip(self, image_paths):
        """
        1) Clear old thumbnails.
        2) Enqueue placeholders + paths.
        3) Kick off _loadNextBatch via QTimer.
        """
        lw = self.listWidget_filmstrip
        lw.clear()

        # Invalidate any previous loader
        self._loadToken += 1
        token = self._loadToken
        self._pendingThumbnails.clear()

        # Create one blank item per image path
        iconSize = lw.iconSize()
        for idx, path in enumerate(image_paths):
            item = QListWidgetItem(QIcon(), "")
            item.setData(Qt.UserRole, idx)
            item.setSizeHint(iconSize)
            lw.addItem(item)
            self._pendingThumbnails.append((item, path, token))

        # Highlight first by default
        if lw.count():
            lw.setCurrentRow(0)

        # Schedule the first batch load
        QTimer.singleShot(self._batchDelay, lambda: self._loadNextBatch(token))

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def _loadNextBatch(self, token):
        """
        Pop up to self._batchSize thumbnails from self._pendingThumbnails,
        assign icons, then reschedule if more remain.
        """
        # Cancel if stale
        if token != self._loadToken:
            return

        lw = self.listWidget_filmstrip
        iconSize = lw.iconSize()

        for _ in range(min(self._batchSize, len(self._pendingThumbnails))):
            item, path, _ = self._pendingThumbnails.pop(0)
            if not os.path.exists(path):
                continue
            pix = QPixmap(path)
            if pix.isNull():
                continue

            thumb = pix.scaled(
                iconSize,
                Qt.KeepAspectRatio,
                Qt.SmoothTransformation
            )
            item.setIcon(QIcon(thumb))

        # More to do?
        if self._pendingThumbnails:
            QTimer.singleShot(self._batchDelay, lambda: self._loadNextBatch(token))

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def browse_ROI_images_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "Select Folder", os.getcwd())
        if not folder:
            return
        self.lineEdit_ROI_images_folder.setText(folder)

        self._on_roi_images_folder_changed()

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def _on_roi_images_folder_changed(self):
        folder = self.lineEdit_ROI_images_folder.text().strip()
        if not folder or not os.path.isdir(folder):
            return

        from GRIME_AI.GRIME_AI_ROI_Analyzer import GRIME_AI_ROI_Analyzer
        temp = GRIME_AI_ROI_Analyzer("", "")
        pairs = temp.generate_file_pairs(folder)
        if not pairs:
            return

        self._pairs = pairs
        self.populate_filmstrip([orig for orig, _ in pairs])

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def analyze_roi(self):
        """
        1) Generate file pairs and populate filmstrip
        2) Run analysis on the first (or clicked) pair
        3) Display composite plot, fallback to original if empty
        4) Overlay top-3 color swatches
        """
        folder = self.lineEdit_ROI_images_folder.text().strip()
        if not folder:
            QMessageBox.warning(self, "ROI Analyzer", "Please specify a folder path.")
            return

        try:
            from GRIME_AI.GRIME_AI_ROI_Analyzer import GRIME_AI_ROI_Analyzer
        except ImportError:
            QMessageBox.warning(self, "ROI Analyzer", "Unable to import ROI Analyzer module.")
            return

        # 1) generate pairs + batched filmstrip population
        temp = GRIME_AI_ROI_Analyzer("", "")
        pairs = temp.generate_file_pairs(folder)
        if not pairs:
            QMessageBox.warning(self, "ROI Analyzer", "No image/mask pairs found.")
            return
        self._pairs = pairs

        # Replace inline loop with batched loader
        image_paths = [orig for orig, _ in pairs]
        self.populate_filmstrip(image_paths)

        # 2) analyze the first pair (index 0) by default
        orig_path, mask_path = pairs[0]
        n_clusters = self.spinBox_numClusters.value()
        analyzer = GRIME_AI_ROI_Analyzer(orig_path, mask_path, clusters=n_clusters)
        analyzer.run_analysis()

        # 3) try to get the composite+metrics pixmap
        try:
            comp_pix = analyzer.get_results_pixmap()
        except Exception as e:
            # fallback: load the original image manually
            img = cv2.imread(orig_path)
            if img is None or img.size == 0:
                QMessageBox.warning(
                    self,
                    "ROI Analyzer",
                    f"Could not generate results pixmap or load original:\n{e}"
                )
                return

            rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
            h, w = rgb.shape[:2]
            comp_pix = QPixmap.fromImage(
                QImage(rgb.data, w, h, w * 3, QImage.Format_RGB888)
            )

        self.label_displayImages.setPixmap(comp_pix)

        # ─── populate metric fields ───
        self.lineEdit_intensity.setText(f"{analyzer.roi_intensity:.2f}")
        self.lineEdit_entropy.setText(f"{analyzer.roi_entropy:.4f}")
        self.lineEdit_Texture.setText(f"{analyzer.roi_texture:.4f}")
        self.lineEdit_GLI.setText(f"{analyzer.mean_gli:.2f}")
        self.lineEdit_GCC.setText(f"{analyzer.mean_gcc:.2f}")

        # 4) overlay the top-3 swatches
        swatches = getattr(analyzer, "dominant_rgb_list", [])[:3]
        self._draw_color_swatches_on_label(swatches, swatch_size=100)

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def extract_ROI_features(self):
        """
        Loop through all (image, mask) pairs in self._pairs, run GRIME_AI_ROI_Analyzer on each,
        export the key metrics to both a CSV and an XLSX file in the folder specified by
        self.lineEdit_ROI_images_folder. In the XLSX, only the filenames appear in the
        image_path and mask_path columns, but those cells are hyperlinked to the full paths.
        """
        # 1) Get and validate output folder path
        output_folder = self.lineEdit_ROI_images_folder.text().strip()
        if not output_folder:
            QMessageBox.warning(self, "No Output Folder", "Please specify an output folder.")
            return

        os.makedirs(output_folder, exist_ok=True)
        csv_path = os.path.join(output_folder, "roi_metrics.csv")
        xlsx_path = os.path.join(output_folder, "roi_metrics.xlsx")

        # 2) Prepare header and container for all rows
        header = [
            "Image Path",
            "Mask Path",
            "ROI Intensity",
            "ROI Entropy",
            "ROI Texture",
            "Mean GLI",
            "Mean GCC",
            "ROI Pixel Count",
            "ROI Area",
            "Image Height",
            "Image Width",
            "Image Total Pixels",
            "ROI Area Percentage"
        ]
        rows = [header]

        # 3) Iterate pairs, run analysis, collect results
        for orig_path, mask_path in self._pairs:
            n_clusters = self.spinBox_numClusters.value()
            from GRIME_AI.GRIME_AI_ROI_Analyzer import GRIME_AI_ROI_Analyzer
            analyzer = GRIME_AI_ROI_Analyzer(orig_path, mask_path, clusters=n_clusters)
            try:
                analyzer.run_analysis()
            except Exception as e:
                print(f"Failed on {orig_path}, {mask_path}: {e}")
                continue

            rows.append([
                orig_path,
                mask_path,
                f"{analyzer.roi_intensity:.2f}",
                f"{analyzer.roi_entropy:.4f}",
                f"{analyzer.roi_texture:.4f}",
                f"{analyzer.mean_gli:.2f}",
                f"{analyzer.mean_gcc:.2f}",
                f"{analyzer.ROI_total_pixels:.2f}",
                f"{analyzer.ROI_total_area:.2f}",
                f"{analyzer.image_height:.2f}",
                f"{analyzer.image_width:.2f}",
                f"{analyzer.image_total_pixels:.2f}",
                f"{analyzer.ROI_percentage:.2f}",
            ])

        # 4) Write out CSV
        with open(csv_path, mode="w", newline="") as csvfile:
            writer = csv.writer(csvfile)
            writer.writerows(rows)

        # 5) Write out XLSX with hyperlinks
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font

            wb = Workbook()
            ws = wb.active
            ws.title = "ROI Metrics"

            # Write header
            for col_idx, title in enumerate(header, start=1):
                ws.cell(row=1, column=col_idx, value=title)

            hyperlink_style = Font(color="0000FF", underline="single")

            # Write data rows
            for row_idx, data in enumerate(rows[1:], start=2):
                orig_full, mask_full, intensity, entropy, texture, gli, gcc, pixel_count, pixel_area, image_height,\
                    image_width, image_total_pixels, roi_area_percentage = data

                # image_path as filename with hyperlink
                img_name = os.path.basename(orig_full)
                cell_img = ws.cell(row=row_idx, column=1, value=img_name)
                cell_img.hyperlink = orig_full
                cell_img.font = hyperlink_style

                # mask_path as filename with hyperlink
                mask_name = os.path.basename(mask_full)
                cell_mask = ws.cell(row=row_idx, column=2, value=mask_name)
                cell_mask.hyperlink = mask_full
                cell_mask.font = hyperlink_style

                # numeric metrics (openpyxl will treat strings as text, so convert back to float)
                ws.cell(row=row_idx, column=3, value=float(intensity))
                ws.cell(row=row_idx, column=4, value=float(entropy))
                ws.cell(row=row_idx, column=5, value=float(texture))
                ws.cell(row=row_idx, column=6, value=float(gli))
                ws.cell(row=row_idx, column=7, value=float(gcc))
                ws.cell(row=row_idx, column=8, value=float(pixel_count))
                ws.cell(row=row_idx, column=9, value=float(pixel_area))
                ws.cell(row=row_idx, column=10, value=float(image_height))
                ws.cell(row=row_idx, column=11, value=float(image_width))
                ws.cell(row=row_idx, column=12, value=float(image_total_pixels))
                ws.cell(row=row_idx, column=13, value=float(roi_area_percentage))
            wb.save(xlsx_path)

        except ImportError:
            QMessageBox.warning(
                self,
                "XLSX Export Skipped",
                "The 'openpyxl' library is not installed. "
                "Install it via 'pip install openpyxl' to enable XLSX export."
            )

        # 6) Notify user
        QMessageBox.information(
            self,
            "Export Complete",
            f"Metrics written to:\n{csv_path}\n"
            + (f"{xlsx_path}" if os.path.exists(xlsx_path) else "(XLSX skipped)")
        )

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def setup_drag_and_drop(self):
        """Configure drag & drop for folder lists and set style for the Segment button."""
        self.listWidget_availableFolders.setDragEnabled(True)
        self.listWidget_availableFolders.setDragDropMode(QAbstractItemView.DragOnly)
        self.listWidget_selectedFolders.setAcceptDrops(True)
        self.listWidget_selectedFolders.setDragDropMode(QAbstractItemView.DropOnly)
        self.listWidget_selectedFolders.installEventFilter(self)
        self.pushButton_Segment.setStyleSheet(
            'QPushButton {background-color: steelblue; color: white; }'
            'QPushButton:disabled {background-color: gray; color: black; }'
        )
        self.pushButton_Segment.setMinimumSize(150, 40)
        self.pushButton_Segment.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def updateSegmentButtonState(self, path: str):
        """Enable the Segment button only if model file and images folder fields are non-empty."""
        """
        Trigger label population when the model path is manually updated.
        """
        path = path.strip()

        if path.lower().endswith('.torch') and os.path.isfile(path):
            self.populate_model_labels(path)
            model_text = self.lineEdit_segmentation_model_file.text().strip()

            images_text = self.lineEdit_segmentation_images_folder.text().strip()

            self.pushButton_Segment.setEnabled(bool(model_text and images_text))

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def updateTrainButtonState(self):
        """Enable the Train button only if at least one selected folder exists."""
        self.pushButton_train.setEnabled(self.listWidget_selectedFolders.count() > 0)

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def populate_segment_images_tab(self):
        """Populate the Segment Images tab with default values."""
        self.lineEdit_segmentation_model_file.setText("")
        self.lineEdit_segmentation_images_folder.setText(os.getcwd())
        self.checkBox_saveModelMasks.setChecked(True)
        self.checkBox_copyOriginalModelImage.setChecked(True)
        print("Segment Images tab populated with default values.")

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def browse_model_training_images_folder(self):
        """Open a dialog to choose a folder and update the folder path field."""
        folder = QFileDialog.getExistingDirectory(self, "Select Folder", os.getcwd())
        folder = os.path.normpath(folder)

        if folder:
            self.lineEdit_model_training_images_path.setText(folder)
            self.populate_available_folders()

            JsonEditor().update_json_entry("Model_Training_Images_Folder", folder)

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def updateCOCOButtonState(self):
        """Enable the Generate COCO button only if a folder is provided in the COCO tab."""
        folder_entered = bool(self.lineEdit_cocoFolder.text().strip())
        self.pushButton_generateCOCO.setEnabled(folder_entered)

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def updateMaskFieldState(self, checked):
        """Enable/disable the mask file field and its Browse button based on the Single Mask checkbox state."""
        self.lineEdit_maskFile.setEnabled(checked)
        self.pushButton_maskBrowse.setEnabled(checked)

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def selectCocoFolder(self):
        """Open a folder chooser for the COCO generation folder."""
        folder = Files().dir()  # You can replace this with QFileDialog.getExistingDirectory if needed.
        if folder:
            self.lineEdit_cocoFolder.setText(folder)

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def selectMaskFile(self):
        """Open a file chooser to select a mask file."""
        mask_file = Files().file()  # Replace with QFileDialog.getOpenFileName if preferred.
        if mask_file:
            self.lineEdit_maskFile.setText(mask_file)

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def getCopyOriginalImage(self):
        return self.checkBox_copyOriginalModelImage.isChecked()

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def getSaveMasks(self):
        return self.checkBox_saveModelMasks.isChecked()

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def getSelectedLabelCategories(self):
        return self.selected_label_categories

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def generateCOCOAnnotations(self):
        """
        Instantiate the CocoGenerator with the folder (and the mask file if Single Mask is checked)
        then generate the COCO annotation JSON in the images folder.
        """
        folder_text = self.lineEdit_cocoFolder.text().strip()
        if not folder_text:
            QMessageBox.warning(self, "COCO Generation", "Please specify a folder path.")
            return

        folder_path = Path(folder_text)
        output_path = folder_path / "instances_default.json"

        if self.checkBox_singleMask.isChecked():
            mask_text = self.lineEdit_maskFile.text().strip()
            if not mask_text:
                QMessageBox.warning(self, "COCO Generation", "Single Mask is selected but no mask file was specified.")
                return
            shared_mask = Path(mask_text)
            if not shared_mask.exists():
                QMessageBox.warning(self, "COCO Generation", f"Specified mask file not found:\n{shared_mask.resolve()}")
                return
            generator = CocoGenerator(folder=folder_path, shared_mask=shared_mask, output_path=output_path)
        else:
            generator = CocoGenerator(folder=folder_path, output_path=output_path)

        generator.generate_annotations()
        QMessageBox.information(self, "COCO Generation",
                                f"COCO annotation file successfully created at:\n{output_path.resolve()}")

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def find_coco_products(self, root_folder):
        """
        Recursively search for all occurrences of 'COCO Products' folders under the given root.
        """
        found = []
        for dirpath, dirnames, _ in os.walk(root_folder):
            print(f"Scanning directory: {dirpath}")
            if "COCO Products" in dirnames:
                coco_path = os.path.join(dirpath, "COCO Products")
                print(f"  Found 'COCO Products' folder: {coco_path}")
                found.append(coco_path)
        if not found:
            print("No 'COCO Products' folder found under", root_folder)
        return found

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def populate_available_folders(self):
        root = Path(self.lineEdit_model_training_images_path.text().strip()).resolve()
        self.listWidget_availableFolders.clear()

        if not root.is_dir():
            QMessageBox.warning(
                self,
                "Invalid Folder",
                f"The selected path is not a directory:\n{root}"
            )
            return

        valid: List[Path] = []
        incomplete: Dict[str, List[str]] = {}

        # ––– NEW: test the root itself
        ok, missing = _check_folder(root)
        print(f"_check_folder on ROOT {root}: ok={ok}, missing={missing}")  # debug
        if ok:
            valid.append(root)
        elif missing:
            incomplete[str(root)] = missing

        # ––– Now recurse into subfolders as before
        for folder in _iter_dirs(root):
            ok, missing = _check_folder(folder)
            print(f"_check_folder on {folder}: ok={ok}, missing={missing}")  # debug
            if ok:
                valid.append(folder)
            elif missing:
                incomplete[str(folder)] = missing

        # ––– Populate or alert “no valid”
        if valid:
            for vf in sorted(set(valid)):
                rel = vf.relative_to(root)
                #JES # Replace "." with actual folder name if it's the root
                #JES display_name = vf.name if str(rel) == "." else str(rel)
                display_name = str(rel)
                self.listWidget_availableFolders.addItem(display_name)
        else:
            QMessageBox.information(
                self,
                "No Valid Training Sets",
                "No folders were found containing a COCO JSON and all its images."
            )

        # ––– Incomplete sets popup
        if incomplete:
            lines = ["Folders missing files:"]
            for fld, miss in incomplete.items():
                lines.append(f"\n{fld}\n  Missing:")
                lines += [f"    • {m}" for m in miss]
            QMessageBox.information(
                self,
                "Incomplete Training Sets",
                "\n".join(lines)
            )

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def move_to_right(self):
        """
        Move selected items from the available folders list to the selected folders list.
        """
        selected_items = self.listWidget_availableFolders.selectedItems()
        for item in selected_items:
            if item:
                if item.text() not in self.transferred_items:
                    self.listWidget_selectedFolders.addItem(item.text())
                    self.transferred_items.add(item.text())
                row = self.listWidget_availableFolders.row(item)
                self.listWidget_availableFolders.takeItem(row)
                print(f"Moved '{item.text()}' from available to selected folders (button).")
        self.updateTrainButtonState()

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def move_to_left(self):
        """
        Move selected items back from the selected folders list to the available folders list,
        then re-sort the available folders.
        """
        selected_items = self.listWidget_selectedFolders.selectedItems()
        for item in selected_items:
            if item:
                row = self.listWidget_selectedFolders.row(item)
                self.listWidget_selectedFolders.takeItem(row)
                if item.text() in self.transferred_items:
                    self.transferred_items.remove(item.text())
                available_items = [self.listWidget_availableFolders.item(i).text() for i in range(self.listWidget_availableFolders.count())]
                available_items.append(item.text())
                available_items.sort()
                self.listWidget_availableFolders.clear()
                for text in available_items:
                    self.listWidget_availableFolders.addItem(text)
                print(f"Moved '{item.text()}' from selected back to available folders (sorted, button).")
        self.updateTrainButtonState()

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def eventFilter(self, source, event):
        """
        Process drag-and-drop events on the selected folders list.
        """
        if source == self.listWidget_selectedFolders:
            if event.type() in (QtCore.QEvent.Type.DragEnter, QtCore.QEvent.Type.DragMove):
                event.accept()
                return True
            elif event.type() == QtCore.QEvent.Type.Drop:
                if event.mimeData().hasText():
                    mime_text = event.mimeData().text()
                    dragged_items = [txt.strip() for txt in mime_text.splitlines() if txt.strip()]
                    for txt in dragged_items:
                        for idx in range(self.listWidget_availableFolders.count()):
                            avail_item = self.listWidget_availableFolders.item(idx)
                            if avail_item.text() == txt:
                                self.listWidget_availableFolders.takeItem(idx)
                                break
                        if txt not in self.transferred_items:
                            self.listWidget_selectedFolders.addItem(txt)
                            self.transferred_items.add(txt)
                            print(f"Dragged '{txt}' from available to selected folders via eventFilter.")
                event.accept()
                self.updateTrainButtonState()
                return True

        if source == self.label_imageAnnotation and event.type() == QtCore.QEvent.Type.MouseButtonDblClick:
            self._openAnnotator()
            return True

        return super().eventFilter(source, event)

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def get_values(self):
        """
        Collect values from dialog controls and return them as a dictionary.
        """
        values = {}
        values["siteName"] = self.lineEdit_siteName.text()
        lr_text = self.lineEdit_learningRates.text()
        try:
            values["learningRates"] = [float(x.strip()) for x in lr_text.split(",") if x.strip()]
        except Exception as e:
            print("Error parsing learning rates:", e)
            values["learningRates"] = lr_text
        values["optimizer"] = self.comboBox_optimizer.currentText()
        values["loss_function"] = self.comboBox_lossFunction.currentText()
        values["weight_decay"] = self.doubleSpinBox_weightDecay.value()
        values["number_of_epochs"] = self.spinBox_epochs.value()
        values["batch_size"] = self.spinBox_batchSize.value()
        values["save_model_frequency"] = self.spinBox_saveFrequency.value()
        values["validation_frequency"] = self.spinBox_validationFrequency.value()
        values["early_stopping"] = self.checkBox_earlyStopping.isChecked()
        values["patience"] = self.spinBox_patience.value()
        values["device"] = self.comboBox_device.currentText()
        values["folder_path"] = self.lineEdit_model_training_images_path.text()
        values["available_folders"] = [self.listWidget_availableFolders.item(i).text() for i in range(self.listWidget_availableFolders.count())]
        values["selected_folders"] = [self.listWidget_selectedFolders.item(i).text() for i in range(self.listWidget_selectedFolders.count())]
        values["segmentation_model_file"] = self.lineEdit_segmentation_model_file.text()
        values["segmentation_images_folder"] = self.lineEdit_segmentation_images_folder.text()
        values["save_model_masks"] = self.checkBox_saveModelMasks.isChecked()
        values["copy_original_model_image"] = self.checkBox_copyOriginalModelImage.isChecked()

        values["num_clusters"] = self.spinBox_numClusters.value()

        return values

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def load_config_from_json(self, filepath):
        """
        Load configuration values from a JSON file.
        """
        with open(filepath, 'r') as f:
            config = json.load(f)
        return config

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def create_custom_json(self):
        """
        Gather all dialog values and create a JSON configuration file.
        """
        from datetime import datetime

        settings_folder = GRIME_AI_Save_Utils().get_settings_folder()
        CONFIG_FILENAME = "site_config.json"
        config_file = os.path.join(settings_folder, CONFIG_FILENAME)
        if os.path.exists(config_file):
            now_str = datetime.today().strftime("%Y_%m_%d_%H_%M_%S")
            backup_filename = os.path.join(settings_folder, f"{now_str}_{CONFIG_FILENAME}")
            os.rename(config_file, backup_filename)
            print(f"Existing {config_file} renamed to {backup_filename}")

        values = self.get_values()
        # EMBED NAME OF LABEL TRAINED ON IN SITE NAME
        values["siteName"] = self.lineEdit_siteName.text()

        seg_images_folder = self.lineEdit_segmentation_images_folder.text().strip()
        if seg_images_folder:
            seg_images_folder = os.path.abspath(seg_images_folder).replace("\\", "/")
        else:
            seg_images_folder = ""

        values["load_model"] = {
            "SAM2_CHECKPOINT": "sam2/checkpoints/sam2.1_hiera_large.pt",
            "MODEL_CFG": "sam2/sam2/configs/sam2.1/sam2.1_hiera_l.yaml",
            "INPUT_DIR": seg_images_folder,
            "OUTPUT_DIR": os.path.join(seg_images_folder, "predictions").replace("\\", "/"),
            "MODEL": "models/Edgewood_19_0.0001_2004_2040.torch"
        }

        seg_model_path = self.lineEdit_segmentation_model_file.text().strip()
        if seg_model_path and seg_model_path.lower().endswith('.torch') and os.path.isfile(seg_model_path):
            abs_model_path = os.path.abspath(seg_model_path.strip()).replace("\\", "/")
            values["load_model"]["MODEL"] = abs_model_path
            print("Updated MODEL path to:", abs_model_path)
        else:
            print("No valid segmentation model file selected; using default MODEL path.")

        root_folder = os.path.abspath(self.lineEdit_model_training_images_path.text().strip()).replace("\\", "/")
        selected_folders = values.get("selected_folders", [])
        if selected_folders:
            new_folders = []
            new_annotations = []
            for folder in selected_folders:
                folder_fwd = folder.replace("\\", "/")
                #JES WE NO LONGER WHAT TO USE THE EXTREMELY LAYERED FOLDER CONVENTION USED BY CVAT
                #JES new_folders.append(root_folder + "/" + folder_fwd + "/images/default")
                #JES new_annotations.append(root_folder + "/" + folder_fwd + "/annotations/instances_default.json")
                new_folders.append(root_folder + "/" + folder_fwd)
                new_annotations.append(root_folder + "/" + folder_fwd + "/instances_default.json")
            values["Path"] = [{
                "siteName": "custom",
                "directoryPaths": {
                    "folders": new_folders,
                    "annotations": new_annotations
                }
            }]
            print("Updated Path section from selected folders.")
        else:
            if hasattr(self, "current_path") and self.current_path:
                values["Path"] = self.current_path
                print("Right listbox is empty; retaining existing Path section from config.")
            else:
                values["Path"] = []
                print("Right listbox is empty and no existing Path data available; setting Path to empty.")

        with open(config_file, "w") as outfile:
            json.dump(values, outfile, indent=4)
        print("Custom JSON file 'site_config.json' created successfully.")

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def initialize_dialog_from_config(self, config):
        self.config = config
        self.setup_from_config_file()

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def accept(self):
        self.create_custom_json()
        super().accept()

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def train(self):
        """
        Called when the Train button is clicked.
        Validates model path, images folder, label selections, and image presence.
        """
        # Label validation if categories were expected
        if self.categories_available:
            selected_labels = []
            for idx in range(self.listWidget_labels.count()):
                item = self.listWidget_labels.item(idx)
                if item and item.isSelected():
                    selected_labels.append(item.text())

        self.create_custom_json()
        print("\nTrain button clicked. Starting training process...")

        settings = self.get_values()
        print("Current settings:", settings)

        self.ml_train_signal.emit()

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def select_segmentation_model(self):
        """
        Open a file dialog to select a segmentation model file (only .torch files).
        Clears the label list and populates it from the model metadata.
        """
        model_file, _ = QFileDialog.getOpenFileName(
            self,
            "Select Segmentation Model",
            "",
            "Torch Model Files (*.torch)"
        )


        if model_file:
            print("Segmentation model selected:", model_file)

            # Clear the label listbox before repopulating
            self.listWidget_labels.clear()

            self.populate_model_labels(model_file)
            self.lineEdit_segmentation_model_file.setText(model_file)
            JsonEditor().update_json_entry("Segmentation_Torch_File", model_file)

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def populate_model_labels(self, model_path):
        """
        Load label categories from side-car JSON or checkpoint metadata.
        Set a fallback flag if none are found.
        """
        labels = None

        '''
        # PROVISIONAL - CURRENTLY, THERE IS NO SIDE-CAR DEPLOYED WITH THE MODEL
        # Try side-car JSON
        meta_json = os.path.splitext(model_path)[0] + ".json"
        if os.path.isfile(meta_json):
            try:
                with open(meta_json, "r") as f:
                    data = json.load(f)
                labels = data.get("categories") or data.get("classes")
            except Exception as e:
                print("Failed to load JSON metadata:", e)
        '''

        # Try torch checkpoint
        if labels is None:
            try:
                ckpt = torch.load(model_path, map_location="cpu")
                labels = ckpt.get("categories") or ckpt.get("meta", {}).get("classes")
            except Exception as e:
                print("Error reading torch file:", e)

        # Clear previous entries
        self.listWidget_labels.clear()

        # No labels found — flag it and allow training
        if not labels:
            QMessageBox.warning(
                self,
                "Load Model Categories",
                "No category list found in:\n" + model_path
            )
            self.categories_available = False
            self.listWidget_labels.addItem("<Older model format>")
            self.listWidget_labels.addItem("<Labels unavailable>")
            self.listWidget_labels.addItem("<ID 2 : Water assumed>")
            self.listWidget_labels.setDisabled(True)
            return

        # If labels exist, reset flag and populate
        self.categories_available = True
        self.listWidget_labels.setDisabled(False)
        for entry in labels:
            if isinstance(entry, dict):
                name = entry.get("name") or entry.get("label") or entry.get("class") or repr(entry)
            else:
                name = str(entry)
            self.listWidget_labels.addItem(name)

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def select_segmentation_images_folder(self):
        """
        Open a folder selection dialog to choose a folder with images for segmentation.
        """
        folder = QFileDialog.getExistingDirectory(self, "Select Images Folder", os.getcwd())
        if folder:
            self.lineEdit_segmentation_images_folder.setText(folder)
            print("Segmentation images folder selected:", folder)

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def segment_images(self):
        """
        Called when the Segment button is clicked.
        Validates label selections, updates configuration, runs segmentation,
        then post-processes outputs according to the two checkboxes.
        """
        if self.categories_available:
            # Ensure that at least one label is selected
            selected_labels = []
            for idx in range(self.listWidget_labels.count()):
                item = self.listWidget_labels.item(idx)
                if item and item.isSelected():
                    selected_labels.append(item.text())

            if not selected_labels:
                QMessageBox.warning(self, "Segmentation Error",
                                    "Please select at least one label before segmenting images.")
                return

            # Populate selected_label_categories from selection
            self.selected_label_categories = []
            if self.categories_available == True:
                for idx in range(self.listWidget_labels.count()):
                    item_text = self.listWidget_labels.item(idx).text().strip()
                    if item_text and item_text in selected_labels:
                        self.selected_label_categories.append({
                            "id": idx + 1,
                            "name": item_text
                        })

            print("Selected categories:", self.selected_label_categories)
        else:   # IF IT IS AN OLDER MODEL NOT CONTAINING LABELS, DEFAULT TO ID: 2, NAME: Water
            self.selected_label_categories = [{"id": 2, "name": "water"}]

        # 1) Gather user settings
        settings = self.get_values()

        # 2) Write JSON config for downstream pipeline
        self.create_custom_json()

        # 3) Kick off the actual segmentation
        self.ml_segment_signal.emit()

        # 4) Close dialog as “Accepted”
        QtCore.QMetaObject.invokeMethod(
            self, 'done', Qt.QueuedConnection,
            QtCore.Q_ARG(int, QDialog.Accepted)
        )

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def resizeEvent(self, event):
        """
        Ensure list widgets update their geometry on dialog resize.
        """
        super().resizeEvent(event)
        self.listWidget_availableFolders.updateGeometry()
        self.listWidget_selectedFolders.updateGeometry()

        if hasattr(self, "_full_canvas"):
            scaled = self._full_canvas.scaled(
                self.label_displayImages.size(),
                Qt.KeepAspectRatio,
                Qt.SmoothTransformation
            )
            self.label_displayImages.setPixmap(scaled)

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def handle_left_item_doubleclick(self, item):
        self.move_items_to_selected([item])

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def handle_right_item_doubleclick(self, item):
        self.move_items_to_available([item])

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def move_items_to_selected(self, items):
        for item in items:
            name = item.text()
            if name not in self.transferred_items:
                self.listWidget_selectedFolders.addItem(name)
                self.transferred_items.add(name)
            for i in range(self.listWidget_availableFolders.count()):
                if self.listWidget_availableFolders.item(i).text() == name:
                    self.listWidget_availableFolders.takeItem(i)
                    break
        self.updateTrainButtonState()

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def move_items_to_available(self, items):
        for item in items:
            name = item.text()
            if name in self.transferred_items:
                self.transferred_items.remove(name)
                available_items = [self.listWidget_availableFolders.item(i).text()
                                   for i in range(self.listWidget_availableFolders.count())]
                available_items.append(name)
                available_items.sort()
                self.listWidget_availableFolders.clear()
                for text in available_items:
                    self.listWidget_availableFolders.addItem(text)
                for i in range(self.listWidget_selectedFolders.count()):
                    if self.listWidget_selectedFolders.item(i).text() == name:
                        self.listWidget_selectedFolders.takeItem(i)
                        break
        self.updateTrainButtonState()

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def reset_lists(self):
        """
        Clear both list widgets and restore available folders from the original list.
        """
        self.listWidget_availableFolders.clear()
        self.listWidget_selectedFolders.clear()
        self.transferred_items.clear()
        for folder in self.original_folders:
            self.listWidget_availableFolders.addItem(QListWidgetItem(folder))
        self.updateTrainButtonState()

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def display_color_swatches(self, rgb_list, swatch_size=100):
        """
        Paints each RGB tuple in rgb_list as a swatch of size swatch_size × swatch_size,
        laid out horizontally, and sets the result into label_displayImages.
        """
        count = len(rgb_list)
        if count == 0:
            return

        # Create a pixmap wide enough for all swatches
        pixmap = QPixmap(swatch_size * count, swatch_size)
        pixmap.fill(Qt.transparent)  # start blank

        painter = QPainter(pixmap)
        for i, (r, g, b) in enumerate(rgb_list):
            color = QColor(r, g, b)
            painter.fillRect(i * swatch_size, 0, swatch_size, swatch_size, color)
        painter.end()

        # Display
        self.label_displayImages.setPixmap(pixmap)

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def _draw_color_swatches_on_label(self, rgb_list, swatch_size=100):
        """
        Builds a canvas containing the current QLabel pixmap (if any) plus
        a horizontal row of color swatches beneath it. The full‐size canvas
        is stored in self._full_canvas, then scaled to fit the label.
        """
        # Determine the base pixmap (the composite + metrics)
        base = self.label_displayImages.pixmap()
        if base:
            base_w, base_h = base.width(), base.height()
        else:
            # no existing image: base area is zero
            base_w, base_h = 0, 0

        # Compute canvas size: width must fit all swatches or base, whichever is wider
        swatch_count = len(rgb_list)
        swatches_w = swatch_count * swatch_size
        total_w = max(base_w, swatches_w)
        total_h = base_h + (swatch_size if swatch_count else 0)

        # Create the full‐size canvas
        canvas = QPixmap(total_w, total_h)
        canvas.fill(Qt.transparent)

        painter = QPainter(canvas)
        # 1) Draw the existing image at the top, if present
        if base:
            painter.drawPixmap(0, 0, base)

        # 2) Draw each swatch beneath the base image
        for i, (r, g, b) in enumerate(rgb_list):
            x = i * swatch_size
            y = base_h
            painter.fillRect(x, y, swatch_size, swatch_size, QColor(r, g, b))
        painter.end()

        # Store the unscaled full‐canvas for future resize events
        self._full_canvas = canvas

        # Scale to fit the QLabel’s current size and set it
        scaled = self._full_canvas.scaled(
            self.label_displayImages.size(),
            Qt.KeepAspectRatio,
            Qt.SmoothTransformation
        )
        self.label_displayImages.setPixmap(scaled)

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def browse_annotation_folder(self):
        """
        Open dialog, select a folder, populate line edit and filmstrip.
        """
        folder = QFileDialog.getExistingDirectory(
            self,
            "Select Annotation Folder",
            os.getcwd()
        )
        if not folder:
            return

        self.lineEdit_annotationFolder.setText(folder)

        labels = self.load_labels_from_annotation(folder)
        self.listWidget_annotationCategories.clear()
        if labels:
            self.listWidget_annotationCategories.addItems(labels)
            print(f"Loaded {len(labels)} label(s) from annotations.json.")

        # Load annotations into internal store

        self.preload_annotation_store_from_coco(folder)

        self.populate_annotation_filmstrip(folder)

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def preload_annotation_store_from_coco(self, folder: str):
        """
        Reads instances_default.json from `folder` and populates self.annotation_store.
        Structure:
            self.annotation_store[full_image_path] = [shape1, shape2, ...]
        Each shape includes 'type', 'points', and 'label'.
        """
        json_path = os.path.join(folder, "instances_default.json")
        if not os.path.exists(json_path):
            print(f"No annotation file found at {json_path}")
            return

        try:
            with open(json_path, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception as e:
            print(f"Error reading {json_path}: {e}")
            return

        image_id_map = {
            img["id"]: os.path.normpath(os.path.join(folder, img["file_name"])).replace("\\", "/")
            # ← 🔧 patched path normalization
            for img in data.get("images", [])
        }
        category_map = {cat["id"]: cat["name"] for cat in data.get("categories", [])}

        self.annotation_store.clear()

        for ann in data.get("annotations", []):
            img_id = ann.get("image_id")
            cat_id = ann.get("category_id")
            segmentation = ann.get("segmentation", [[]])[0]

            if img_id not in image_id_map or cat_id not in category_map or not segmentation:
                continue

            img_path = image_id_map[img_id]
            label = {"id": cat_id, "name": category_map[cat_id]}

            pts = []
            for i in range(0, len(segmentation), 2):
                x, y = segmentation[i], segmentation[i + 1]
                pts.append(QtCore.QPointF(x, y))

            shape = {
                "type": "drag",  # ← 🔧 changed type from "polygon" to "drag"
                "points": pts,
                "label": label
            }

            if img_path not in self.annotation_store:
                self.annotation_store[img_path] = []

            self.annotation_store[img_path].append(shape)

        print(f"Preloaded {len(self.annotation_store)} images from instances_default.json")

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def populate_annotation_filmstrip(self, folder):
        """
        List all supported images in `folder`, fill the filmstrip.
        """
        self.listWidget_annotationFilmstrip.clear()

        # gather all image files
        exts = ('.png', '.jpg', '.jpeg', '.bmp', '.tif', '.tiff')
        paths = sorted(
            Path(folder).glob('*')
        )
        images = [str(p) for p in paths if p.suffix.lower() in exts]
        self._annotation_image_paths = images

        # preload existing .anno.json for each image
        for img_path in images:
            self.load_annotations_for_image(img_path)

        # update the category list based on preloaded data
        self._refresh_annotation_categories()         # NEW: populate categories

        icon_size = self.listWidget_annotationFilmstrip.iconSize()
        for idx, img_path in enumerate(images):
            pix = QPixmap(img_path)
            if pix.isNull():
                continue
            thumb = pix.scaled(icon_size, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            item = QListWidgetItem(QIcon(thumb), "")
            # store index so we can look it up on click
            item.setData(Qt.UserRole, idx)
            self.listWidget_annotationFilmstrip.addItem(item)

        # auto-select first
        if self.listWidget_annotationFilmstrip.count():
            self.listWidget_annotationFilmstrip.setCurrentRow(0)
            # display it immediately
            self.display_annotation_image(self.listWidget_annotationFilmstrip.currentItem())

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def display_annotation_image(self, item):
        """
        Show the full image corresponding to the clicked thumbnail.
        """
        idx = item.data(Qt.UserRole)
        path = self._annotation_image_paths[idx]
        pix = QPixmap(path)
        if pix.isNull():
            return

        # scale to fit the label, keep aspect ratio
        lbl = self.label_imageAnnotation
        scaled = pix.scaled(
            lbl.size(),
            Qt.KeepAspectRatio,
            Qt.SmoothTransformation
        )
        lbl.setPixmap(scaled)

    # ------------------------------------------------------------------------------------------------------------------
    #
    # ------------------------------------------------------------------------------------------------------------------
    def add_annotation_category(self):
        """
        Called when the user presses Enter in the annotation combo's line edit.
        Adds the current text as a new category if it’s non-empty and not already present.
        """
        text = self.comboBox_annotationCategories.currentText().strip()
        if not text:
            return

        # Only add unique, non-empty entries
        if self.comboBox_annotationCategories.findText(text) == -1:
            self.comboBox_annotationCategories.addItem(text)

        # Clear the editor so it's ready for the next entry
        self.comboBox_annotationCategories.lineEdit().clear()

# End of GRIME_AI_ML_ImageProcessingDlg class.
