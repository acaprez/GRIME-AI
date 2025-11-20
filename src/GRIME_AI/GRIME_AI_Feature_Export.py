#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# Author: John Edward Stranzl, Jr.
# Affiliation(s): University of Nebraska-Lincoln, Blade Vision Systems, LLC
# Contact: jstranzl2@huskers.unl.edu, johnstranzl@gmail.com
# Created: Mar 6, 2022
# License: Apache License, Version 2.0, http://www.apache.org/licenses/LICENSE-2.0

import os
import datetime
import cv2
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage

from GRIME_AI.GRIME_AI_QProgressWheel import QProgressWheel
from GRIME_AI.GRIME_AI_Utils import GRIME_AI_Utils
from GRIME_AI.GRIME_AI_TimeStamp_Utils import GRIME_AI_TimeStamp_Utils
from GRIME_AI.GRIME_AI_Color import GRIME_AI_Color
from GRIME_AI.GRIME_AI_Vegetation_Indices import GRIME_AI_Vegetation_Indices
from GRIME_AI.GRIME_AI_roiData import GRIME_AI_roiData

from GRIME_AI.GRIME_AI_Texture import GLCMTexture, LBPTexture, GaborTexture, WaveletTexture, FourierTexture

# ======================================================================================================================
#
# ======================================================================================================================
class GRIME_AI_Feature_Export:
    def __init__(self):
        csvFilename = ''
        self.instance = 1

    # ==================================================================================================================
    #
    # ==================================================================================================================
    def create_training_data_filename(self, imageFileFolder):
        csvFilename = 'TrainingData_' + datetime.datetime.now().strftime("%Y%m%d_%H%M%S") + '.csv'
        imageQualityFile = os.path.join(imageFileFolder, csvFilename)
        csvFilename = open(imageQualityFile, 'a', newline='')

        xlsxFilename = 'TextureData_' + datetime.datetime.now().strftime("%Y%m%d_%H%M%S") + '.xlsx'
        xlsxFilename = os.path.join(imageFileFolder, xlsxFilename)

        return csvFilename, xlsxFilename

    # ==================================================================================================================
    #
    # ==================================================================================================================
    def ExtractFeatures(self, imagesList, imageFileFolder, roiList, colorSegmentationParams, greenness_index_list):

        # ----------------------------------------------------------------------------------------------------------
        # CREATE PROGRESS WHEEL
        # ----------------------------------------------------------------------------------------------------------
        myGRIMe_Color = GRIME_AI_Color()

        # GENERATE LIST OF IMAGE FILES IN FOLDER
        videoFileList = imagesList
        nFrameCount = len(videoFileList)

        progressBar = QProgressWheel(0, (len(roiList) * len(videoFileList)) + 1)
        progressBar.show()
        progressBarIndex = 1

        if nFrameCount > 0:
            # CREATE TRAINING DATA FILENAME
            csvFile, xlsFile = self.create_training_data_filename(imageFileFolder)

            # CREATE THE OUTPUT FILE COLUMN HEADER
            nMaxNumColorClusters = GRIME_AI_Utils().getMaxNumColorClusters(roiList)

            '''
            for roiObj in roiList:

                # Build the initial columns
                header_cols = ["Training Filename", "Date (ISO)", "Time (ISO)",  "Intensity", "Entropy"]
                base_header = ",".join(header_cols)

                for greenness in greenness_index_list:
                    base_header = base_header + ',' + greenness.get_name()

                base_header = base_header + ',,'

                # Create the template for HSV headings based on the ROI name
                template = f", {roiObj.getROIName()}: #"

                # Build the HSV part of the header
                if nMaxNumColorClusters == 1:
                    hsv_cols = ''.join(template.replace('#', char) for char in ['H', 'S', 'V'])
                else:
                    hsv_cols = ''.join(
                        ''.join(f"{template.replace('#', char)}_{idx}" for idx in range(nMaxNumColorClusters))
                        for char in ['H', 'S', 'V']
                    )

                # Combine all parts and add a newline at the end
                strOutputString = base_header + hsv_cols + "\n"
                strOutputString += "\n"

                csvFile.write(strOutputString)

                # LOAD THE IMAGE FILE
                img = myGRIMe_Color.loadColorImage(roiObj.getTrainingImageName())

                # CONVERT TO GRAY SCALE
                gray = cv2.cvtColor(img, cv2.COLOR_RGB2GRAY)
                height, width = gray.shape
                gray = cv2.resize(gray, (width, height))

                texture = -999

                # CREATE HYPERLINK TO FILE
                strOutputString = self.create_hyperlink(roiObj.getTrainingImageName())
                strOutputString = strOutputString + ','     # DATE
                strOutputString = strOutputString + ','     # TIME

                # COMPUTE INTENSITY FOR THE ENTIRE IMAGE
                # ------------------------------------------------------------------------------------------------------
                intensity = cv2.mean(gray)[0]  # The range for a pixel's value in grayscale is (0-255), 127 lies midway
                strOutputString = strOutputString + ',' + '%3.4f' % intensity

                # COMPUTE ENTROPY FOR THE ENTIRE IMAGE
                # ------------------------------------------------------------------------------------------------------
                entropyValue = self.calcEntropy(gray)
                strOutputString = strOutputString + ',' + '%3.4f' % entropyValue

                # CALCULATE THE GREENNESS INDEX FOR THE ROI
                # ------------------------------------------------------------------------------------------------------
                try:
                    for greenness in greenness_index_list:
                        greenness_updated = GRIME_AI_Vegetation_Indices().get_greenness(greenness, img)
                        strOutputString = strOutputString + ',' + '%3.4f' % greenness_updated.get_value()
                except:
                    pass

                # SKIP TWO COLUMNS TO ALIGN THE WHOLE IMAGE HEADER WITH THE ROI HEADERS
                strOutputString = strOutputString + ',,'

                hsvClusterCenters, hist = roiObj.getHSVClusterCenters()
                for i in range(nMaxNumColorClusters):
                    strOutputString = strOutputString + ',' + str(hsvClusterCenters[i][0]) + ',' + str(hsvClusterCenters[i][1]) + ',' + str(hsvClusterCenters[i][2])
                strOutputString = strOutputString + '\n'

                csvFile.write(strOutputString)
            '''


            self.extract_ROI_features(csvFile, xlsFile, videoFileList, roiList, colorSegmentationParams, greenness_index_list, progressBarIndex, progressBar)


            csvFile.close()

            # CLOSE AND DELETE THE PROGRESSBAR
            progressBar.close()
            del progressBar

    # ==================================================================================================================
    #
    # ==================================================================================================================
    def build_scalar_header(self, nClusters, roiList, colorSegmentationParams, greenness_index_list):
        # CREATE HEADER FOR THE ATTIRBUTE OF THE ENTIRE IMAGE
        header = 'Image, Date (ISO), Time (ISO)'

        #if colorSegmentationParams.wholeImage:
        if 1:
            header = self.build_image_scalar_header(header, roiList, colorSegmentationParams)

        if colorSegmentationParams.ROI:
            header = self.build_ROI_scalar_header(header, roiList, colorSegmentationParams, greenness_index_list)

        return header



    # ==================================================================================================================
    #
    # ==================================================================================================================
    def extract_ROI_features(self, csvFile, xlsFile, videoFileList, roiList, colorSegmentationParams, greenness_index_list, progressBarIndex, progressBar):

        myGRIMe_Color = GRIME_AI_Color()

        # ----------------------------------------------------------------------------------------------------------
        # PROCESS IMAGES
        # ----------------------------------------------------------------------------------------------------------
        # E:\\000 - University of Nebraska\\2022_USGS104b\\imagery\\PBT_MittelstetsMeanderCam\\20220709

        # ONCE THE NUMBER OF COLOR CLUSTERS HAS BEEN SELECTED AND AN ROI TRAINED, LOCK THE NUMBER OF COLOR CLUSTERS
        # AND TRAIN ALL SUBSEQUENT ROIs FOR THE SAME NUMBER OF COLOR CLUSTERS
        nClusters = colorSegmentationParams.numColorClusters

        header = self.build_scalar_header(nClusters, roiList, colorSegmentationParams, greenness_index_list)

        # WRITE THE HEADER TO THE CSV
        csvFile.write(header)

        try:
            for fname in videoFileList:

                if os.path.isfile(fname.fullPathAndFilename):

                    # CREATE A STRING THAT IS A HYPERLINK TO THE FILE
                    strHyperlink = self.create_hyperlink(fname.fullPathAndFilename)

                    # EXTRACT DATE/TIME STAMP FROM IMAGE
                    myGRIME_AI_TimeStamp_Utils = GRIME_AI_TimeStamp_Utils()
                    myGRIME_AI_TimeStamp_Utils.detectDateTime(fname.fullPathAndFilename)
                    strDate, strTime = myGRIME_AI_TimeStamp_Utils.extractDateTime(fname.fullPathAndFilename)

                    # WRITE THE HYPERLINK, DATE, AND TIME STAMP TO THE OUTPUT FILE
                    strOutputString = '%s,%s,%s' % (strHyperlink, strDate, strTime)

                    # LOAD THE IMAGE FILE TO START THE PROCESS OF EXTRACTING FEATURE DATA
                    img = myGRIMe_Color.loadColorImage(fname.fullPathAndFilename)

                    # if colorSegmentationParams.wholeImage:
                    if 1:
                        # BLUR THE IMAGE
                        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

                        # EXTRACT 'n' DOMINANT HSV COLORS
                        hist, clusterCenters = myGRIMe_Color.extractDominant_HSV(img, nClusters)

                        if colorSegmentationParams.Intensity:
                            # IMAGE INTENSITY CALCULATIONS
                            # The range for a pixel's value in grayscale is (0-255), 127 lies midway
                            strOutputString = strOutputString + ', %3.4f' % cv2.mean(gray)[0]

                        if colorSegmentationParams.ShannonEntropy:
                            # COMPUTE ENTROPY FOR ENTIRE IMAGE
                            strOutputString = strOutputString + ', %3.4f' % self.calcEntropy(gray)

                        if colorSegmentationParams.Texture:
                            texture = -999
                            strOutputString = strOutputString + ', %3.2f' % texture

                        try:
                            for greenness in greenness_index_list:
                                greenness_updated = GRIME_AI_Vegetation_Indices().get_greenness(greenness, img)
                                strOutputString = strOutputString + ',' + '%3.4f' % greenness_updated.get_value()
                        except ValueError:
                            pass

                        if colorSegmentationParams.HSV:
                            # CONVERT FROM OpenCV's HSV HUE DATA FORMAT 0 to 180 DEGREES TO THE HSV STANDARD FORMAT OF 0 to 360 DEGREES
                            # Concatenate the HSV values for each cluster succinctly
                            strOutputString += ''.join(
                                f", {float(center[0]):3.2f}, {float(center[1]):3.2f}, {float(center[2]):3.2f}"
                                for center in clusterCenters[:nClusters]
                            )

                    # --------------------------------------------------------------------------------------------------
                    #
                    # --------------------------------------------------------------------------------------------------
                    progressBar.setValue(progressBarIndex)
                    progressBarIndex = progressBarIndex + 1

                    if colorSegmentationParams.ROI:
                        strROI = self.calculate_ROI_scalars(img, roiList, colorSegmentationParams, greenness_index_list,
                                                            xlsFile)

                        strOutputString = strOutputString + strROI

                    # WRITE STRING TO CSV FILE
                    strOutputString = strOutputString + '\n'
                    csvFile.write(strOutputString)
        except ValueError:
            exceptValue = 0



    # ==================================================================================================================
    #
    # ==================================================================================================================
    def build_image_scalar_header(self, header, roiList, colorSegmentationParams):
        nClusters = GRIME_AI_Utils().getMaxNumColorClusters(roiList)

        if colorSegmentationParams.Intensity:
            header = header + ", Intensity"

        if colorSegmentationParams.ShannonEntropy:
            header = header + ", Entropy"

        if colorSegmentationParams.Texture:
            header = header + ", Texture"

        if colorSegmentationParams.GCC:
            header = header + ", GCC"

        if colorSegmentationParams.GLI:
            header = header + ", GLI"

        if colorSegmentationParams.NDVI:
            header = header + ", NDVI"

        if colorSegmentationParams.ExG:
            header = header + ", ExG"

        if colorSegmentationParams.RGI:
            header = header + ", RGI"

        if colorSegmentationParams.HSV:
            template = ', Image_#'

            if nClusters == 1:
                header += template.replace('#', 'H') + template.replace('#', 'S') + template.replace('#', 'V')
            else:
                header += ''.join(
                    f"{template.replace('#', channel)}: {idx}" for idx in range(nClusters) for channel in
                    ['H', 'S', 'V']
                )

        return header


    # ==================================================================================================================
    #
    # ==================================================================================================================
    def build_ROI_scalar_header(self, header, roiList, colorSegmentationParams, greenness_index_list):

        newHeader = header

        # ADD HEADER TO CSV FILE FOR EACH ROI ASSUMING ROIs HAVE BEEN CREATED
        for roiObj in roiList:
            template = ', ' + (roiObj.getROIName() + ': ') + '#'

            if colorSegmentationParams.Intensity:
                newHeader = (newHeader + template).replace('#', 'Intensity')

            if colorSegmentationParams.ShannonEntropy:
                newHeader = (newHeader + template).replace('#', 'Entropy')

            if colorSegmentationParams.Texture:
                newHeader = (newHeader + template).replace('#', 'Texture')

            if colorSegmentationParams.GCC:
                newHeader = (newHeader + template).replace('#', 'GCC')

            if colorSegmentationParams.GLI:
                newHeader = (newHeader + template).replace('#', 'GLI')

            if colorSegmentationParams.NDVI:
                newHeader = (newHeader + template).replace('#', 'NDVI')

            if colorSegmentationParams.ExG:
                newHeader = (newHeader + template).replace('#', 'ExG')

            if colorSegmentationParams.RGI:
                newHeader = (newHeader + template).replace('#', 'RGI')

            # IF THERE IS MORE THAN ONE (1) ROI, APPEND AN INDEX ONTO THE HEADER LABEL

            nClusters = roiObj.getNumColorClusters()

            if colorSegmentationParams.HSV:
                template = f", {roiObj.getROIName()}: "

                if nClusters == 1:
                    channels = ['H', 'S', 'V']
                    for channel in channels:
                        newHeader += f"{template}{channel}"
                else:
                    channels = ['H', 'S', 'V']
                    for idx in range(nClusters):
                        for channel in channels:
                            newHeader += f"{template}{channel}_{idx}"

        newHeader = newHeader + '\n'

        return newHeader

    # ==================================================================================================================
    # CALCULATE THE FEATURE SCALARS FOR THE VARIOUS ROIs AND SAVE THEM TO THE CSV FILE
    # ==================================================================================================================
    def calculate_ROI_scalars(self, img, roiList, colorSegmentationParams, greenness_index_list, xlsFile):

        strOutputString = ''

        for roiObj in roiList:
            texture = -999

            if 0:
                self.extract_texture_and_savCe(
                    image=img,
                    excel_filename=xlsFile,
                    plot_folder="plots",
                    file_name="example_image",
                    plot_option="embed"  # Options: "file", "embed", or "both"
                )

            nClusters = roiObj.getNumColorClusters()

            rgb = GRIME_AI_roiData().extractROI(roiObj.getImageROI(), img)

            # CONVERT THE IMAGE FROM BGR TO RGB AND HSV
            try:
                hsv = cv2.cvtColor(rgb, cv2.COLOR_RGB2HSV)
                gray = cv2.cvtColor(rgb, cv2.COLOR_RGB2GRAY)
            except Exception:
                tempValue = 0

            # IMAGE INTENSITY CALCULATIONS
            intensity = cv2.mean(gray)[0]  # The range for a pixel's value in grayscale is (0-255), 127 lies midway

            # COMPUTE ENTROPY FOR ROI
            entropyValue = self.calcEntropy(gray)

            # EXTRACT 'n' DOMINANT HSV COLORS
            myGRIMe_Color = GRIME_AI_Color()
            qhsvImg, hsvClusterCenters, hist = myGRIMe_Color.KMeans(hsv, nClusters)

            # KMeans QUANTIZES THE HUE VALUE TO 0..180 WHEN THE ACTUAL HSV COLOR SPACE HUE VALUE 0..360.
            # THEREFORE WE MULTIPLY THE KMeans HUE VALUE BY 2 TO STANDARDIZE ON THE ACTUAL COLOR SPACE HUE RANGE.
            hsvClusterCenters[:, 0] = hsvClusterCenters[:, 0] * 2.0

            if colorSegmentationParams.Intensity:
                strOutputString = strOutputString + ', %3.4f' % intensity
            if colorSegmentationParams.ShannonEntropy:
                strOutputString = strOutputString + ', %3.4f' % entropyValue
            if colorSegmentationParams.Texture:
                strOutputString = strOutputString + ', %3.2f' % texture

            try:
                for greenness in greenness_index_list:
                    greenness_updated = GRIME_AI_Vegetation_Indices().get_greenness(greenness, rgb)
                    strOutputString = strOutputString + ',' + '%3.4f' % greenness_updated.get_value()
            except:
                pass

            if colorSegmentationParams.HSV:
                for idx in range(nClusters):
                    # CONVERT FROM OpenCV's HSV HUE DATA FORMAT 0 to 180 DEGREES TO THE HSV STANDARD FORMAT OF 0 to 360 DEGREES
                    h = float(hsvClusterCenters[idx][0])
                    s = float(hsvClusterCenters[idx][1])
                    v = float(hsvClusterCenters[idx][2])
                    strOutputString = strOutputString + ', %3.2f, %3.2f, %3.2f' % (h, s, v)

        return strOutputString


    # ==================================================================================================================
    #
    # ==================================================================================================================
    def calcEntropy(self, img):
        entropy = []

        hist = cv2.calcHist([img], [0], None, [256], [0, 255])
        total_pixel = img.shape[0] * img.shape[1]

        for item in hist:
            probability = item / total_pixel
            if probability == 0:
                en = 0
            else:
                en = -1 * probability * (np.log(probability) / np.log(2))
            entropy.append(en)

        try:
            sum_en = sum(entropy)
        except Exception:
            sum_en = 0.0

        return sum_en

    def create_hyperlink(self, file_path: str) -> str:
        """
        Returns an Excel HYPERLINK formula for the given file path.
        This function cleans the file path and escapes it appropriately for CSV.
        """
        # Clean the file path by stripping extra whitespace and any extraneous quotes
        clean_path = file_path.strip().strip('"')

        # Create the hyperlink formula
        formula = f'=HYPERLINK("{clean_path}", "{os.path.basename(clean_path)}")'

        # If the formula contains commas, enclose the entire field in extra quotes and escape inner quotes.
        if ',' in formula:
            safe_formula = formula.replace('"', '""')
            formula = f'"{safe_formula}"'

        return formula


    def extract_texture_and_save(self, image, excel_filename, plot_folder, file_name, plot_option="both"):
        """
        Process an image, extract texture features, save them to an Excel file, and handle Fourier profile plots.

        Parameters:
            image (numpy array): Input image data.
            excel_filename (str): Name of the Excel file to append texture features.
            plot_folder (str): Folder path to save Fourier profile plots (if required).
            file_name (str): Name of the image (used for labeling data in Excel).
            plot_option (str): Options for handling the plot:
                               - "file": Save the plot as an image file.
                               - "embed": Embed the plot into the Excel file.
                               - "both": Save the plot as an image file and embed it into the Excel file.
        """
        #if not os.path.exists(excel_filename):
        #    with open(excel_filename, 'w') as file:
        #        pass  # Create an empty file
        #    print(f"File '{excel_filename}' created.")
        # Ensure the file exists
        if not os.path.exists(excel_filename):
            # Create the file with a default visible sheet
            with pd.ExcelWriter(excel_filename, engine="openpyxl") as writer:
                # Add a default empty DataFrame to create a visible sheet
                pd.DataFrame().to_excel(writer, sheet_name="Default_Sheet", index=False)

        # Ensure the output folder for plots exists if saving as files
        if plot_option in ["file", "both"]:
            os.makedirs(plot_folder, exist_ok=True)

        # Instantiate texture measurement objects
        glcm_obj = GLCMTexture()
        gabor_obj = GaborTexture()
        lbp_obj = LBPTexture()
        wavelet_obj = WaveletTexture()
        fourier_obj = FourierTexture()

        # Compute features using each method
        glcm_features = glcm_obj.compute_features(image)
        gabor_features = gabor_obj.compute_features(image)
        lbp_features = lbp_obj.compute_features(image)
        wavelet_features = wavelet_obj.compute_features(image)
        fourier_features = fourier_obj.compute_features(image)

        # Create DataFrames for the texture features
        glcm_df = pd.DataFrame({"GLCM Features": glcm_features})
        gabor_df = pd.DataFrame({"Gabor Features": gabor_features})
        lbp_df = pd.DataFrame({"LBP Histogram": lbp_features})
        wavelet_df = pd.DataFrame({"Wavelet Features": wavelet_features})
        fourier_df = pd.DataFrame({"Fourier Radial Profile": fourier_features})

        # Write DataFrames to the Excel file under separate sheets for the image
        try:
            with pd.ExcelWriter(excel_filename, mode="a", engine="openpyxl") as writer:
                glcm_df.to_excel(writer, sheet_name=f"{file_name}_GLCM")
                gabor_df.to_excel(writer, sheet_name=f"{file_name}_Gabor")
                lbp_df.to_excel(writer, sheet_name=f"{file_name}_LBP")
                wavelet_df.to_excel(writer, sheet_name=f"{file_name}_Wavelet")
                fourier_df.to_excel(writer, sheet_name=f"{file_name}_Fourier")
        except Exception:
            pass

        # Plot the F    self._book = load_workbook(self._handles.handle, **engine_kwargs)ourier radial profile
        plt.figure(figsize=(6, 4))
        plt.plot(fourier_features, marker='o')
        plt.title(f"Fourier Radial Profile - {file_name}")
        plt.xlabel("Radial Bin")
        plt.ylabel("Average Magnitude")
        plt.grid(True)

        # Handle plot based on user option
        plot_filename = os.path.join(plot_folder, f"{file_name}_fourier_profile.png")
        if plot_option in ["file", "both"]:
            plt.savefig(plot_filename)  # Save the plot to a file
        if plot_option in ["embed", "both"]:
            plt.savefig("temp_plot.png")  # Temporary file to embed the plot
            plt.close()

            # Open the workbook to embed the plot
            workbook = load_workbook(excel_filename)
            worksheet = workbook[f"{file_name}_Fourier"]
            img = ExcelImage("temp_plot.png")
            worksheet.add_image(img, "B10")  # Embed image at cell B10
            workbook.save(excel_filename)

        print(
            f"Texture features for {file_name} written to {excel_filename}. Plots handled as per option '{plot_option}'.")
