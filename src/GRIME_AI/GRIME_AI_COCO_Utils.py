#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# Author: John Edward Stranzl, Jr.
# Affiliation(s): University of Nebraska-Lincoln, Blade Vision Systems, LLC
# Contact: jstranzl2@huskers.unl.edu, johnstranzl@gmail.com
# Created: Mar 6, 2022
# License: Apache License, Version 2.0, http://www.apache.org/licenses/LICENSE-2.0

#!/usr/bin/env python3
"""
coco_utils.py

Utility class for validating, cleaning, and analyzing COCO 1.0 JSON.
"""

import json
import os
import sys
import matplotlib.pyplot as plt
from PyQt5.QtWidgets import QApplication

import openpyxl
from openpyxl.utils import get_column_letter

from collections import Counter
import pandas as pd

from typing import Any, Dict, List, Tuple
import cv2

import numpy as np
from pycocotools.coco import COCO
from pycocotools import mask as maskUtils

from GRIME_AI_QProgressWheel import QProgressWheel


# ======================================================================================================================
# ======================================================================================================================
# ===   ===   ===   ===   ===   ===   ===   === class GRIME_AI_COCO_Utils  ===   ===   ===   ===   ===   ===   ===   ===
# ======================================================================================================================
# ======================================================================================================================
class GRIME_AI_COCO_Utils:
    """Utility class for validating, cleaning, and summarizing a COCO-format JSON file."""

    def __init__(self, folder_path: str) -> None:
        self.folder_path: str = folder_path
        self.json_filename: str = ""
        self.json_path: str = ""
        self.data: Dict[str, Any] = {}


    # ------------------------------------------------------------------------------------------------------------------
    # ------------------------------------------------------------------------------------------------------------------
    def find_json_file(self) -> None:
        try:
            entries = os.listdir(self.folder_path)
        except FileNotFoundError:
            print(f"Error: folder not found -> {self.folder_path}")
            sys.exit(1)

        json_files = [f for f in entries if f.lower().endswith(".json")]
        if not json_files:
            print(f"No JSON file found in {self.folder_path}")
            sys.exit(1)

        if len(json_files) > 1:
            print(f"Multiple JSON files found: {json_files}")
            sys.exit(1)

        self.json_filename = json_files[0]
        self.json_path = os.path.join(self.folder_path, self.json_filename)

    # ------------------------------------------------------------------------------------------------------------------
    # ------------------------------------------------------------------------------------------------------------------
    def load_json(self) -> None:
        with open(self.json_path, "r", encoding="utf-8") as f:
            self.data = json.load(f)
        if "images" not in self.data:
            print("Invalid COCO JSON: missing 'images' section.")
            sys.exit(1)

    # ------------------------------------------------------------------------------------------------------------------
    # ------------------------------------------------------------------------------------------------------------------
    def load_coco(self) -> None:
        """
        Convenience method to find and load the COCO JSON in one call.
        """
        self.find_json_file()
        self.load_json()

    # ------------------------------------------------------------------------------------------------------------------
    # ------------------------------------------------------------------------------------------------------------------
    def check_images(
        self,
    ) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
        present_images: List[Dict[str, Any]] = []
        missing_images: List[Dict[str, Any]] = []

        for img in self.data["images"]:
            fname = img.get("file_name", "")
            path = os.path.join(self.folder_path, fname)
            if fname and os.path.isfile(path):
                present_images.append(img)
            else:
                missing_images.append(img)

        return present_images, missing_images

    # ------------------------------------------------------------------------------------------------------------------
    # ------------------------------------------------------------------------------------------------------------------
    def backup_original(self) -> None:
        orig = os.path.join(self.folder_path, self.json_filename)
        backup = orig + ".ORIGINAL"
        os.rename(orig, backup)

    # ------------------------------------------------------------------------------------------------------------------
    # ------------------------------------------------------------------------------------------------------------------
    def clean_data(
        self, present_images: List[Dict[str, Any]]
    ) -> Dict[str, Any]:
        new_data = dict(self.data)
        new_data["images"] = present_images

        if "annotations" in self.data:
            valid_ids = {img["id"] for img in present_images}
            new_data["annotations"] = [
                ann
                for ann in self.data["annotations"]
                if ann.get("image_id") in valid_ids
            ]

        return new_data

    # ------------------------------------------------------------------------------------------------------------------
    # ------------------------------------------------------------------------------------------------------------------
    def write_json(self, new_data: Dict[str, Any]) -> None:
        with open(self.json_path, "w", encoding="utf-8") as f:
            json.dump(new_data, f, indent=4)

    # ------------------------------------------------------------------------------------------------------------------
    # ------------------------------------------------------------------------------------------------------------------
    def process(self) -> None:
        self.find_json_file()
        self.load_json()

        present, missing = self.check_images()
        if not missing:
            print("All images in JSON are present.")
            return

        self.backup_original()
        cleaned = self.clean_data(present)
        self.write_json(cleaned)
        print("New JSON created for the images available in the folder.")

    # ------------------------------------------------------------------------------------------------------------------
    # ------------------------------------------------------------------------------------------------------------------
    def get_image_label_counts(self) -> Dict[str, Dict[str, int]]:
        """
        Returns a mapping from image file names to a dict of category names
        and their respective annotation counts (mask counts).
        """
        # Build lookups
        id_to_name = {c["id"]: c["name"] for c in self.data.get("categories", [])}
        id_to_file = {img["id"]: img["file_name"] for img in self.data.get("images", [])}

        # Initialize counts dict
        counts: Dict[str, Dict[str, int]] = {}
        for fname in id_to_file.values():
            counts[fname] = {name: 0 for name in id_to_name.values()}

        # Tally annotations
        for ann in self.data.get("annotations", []):
            img_id = ann.get("image_id")
            cat_id = ann.get("category_id")
            if img_id in id_to_file and cat_id in id_to_name:
                fname = id_to_file[img_id]
                cat_name = id_to_name[cat_id]
                counts[fname][cat_name] += 1

        return counts

    # ------------------------------------------------------------------------------------------------------------------
    # ------------------------------------------------------------------------------------------------------------------
    def write_image_label_counts_to_xlsx(self, output_file: str) -> None:
        """
        Writes the image-wise label counts to an Excel file.

        The sheet will have one row per image, with columns:
        [Image Name, <Label1>, <Label2>, …].

        :param output_filepath: Path to write the .xlsx file.
        """
        # 1) Ensure data is loaded
        if not self.data:
            raise RuntimeError("COCO data not loaded. Call load_coco() first.")

        # 2) Fetch counts dict: { image_name: { label: count, … }, … }
        counts = self.get_image_label_counts()

        # 3) Create workbook & sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Image Label Counts"

        # 4) Build header row
        #    - first col is "Image"
        #    - rest are all label names, sorted alphabetically
        all_labels = list(next(iter(counts.values())).keys())
        header = ["Image"] + sorted(all_labels)
        ws.append(header)

        # 5) Populate rows
        for img_name, label_map in counts.items():
            row = [img_name] + [label_map.get(lbl, 0) for lbl in header[1:]]
            ws.append(row)

        # 6) Auto-adjust column widths
        for idx, col_title in enumerate(header, start=1):
            max_length = max(
                len(str(cell.value)) for cell in ws[get_column_letter(idx)]
            )
            # add a little padding
            ws.column_dimensions[get_column_letter(idx)].width = max_length + 2

        # 7) Save to disk
        output_file = os.path.join(self.folder_path, output_file)
        wb.save(output_file)

    # ==================================================================================================================
    # ==================================================================================================================
    # ==================================================================================================================
    def create_overlay(self, image, color_mask, output_dir, base, ext=".png"):
        """
        Overlay the color mask transparently over the original image and save with '_overlay' suffix.
        """
        alpha = 0.5  # transparency factor
        overlay = cv2.addWeighted(image, 1.0, color_mask, alpha, 0)

        out_path = os.path.join(output_dir, f"{base}_overlay{ext}")
        cv2.imwrite(out_path, overlay)
        print(f"Overlay saved: {out_path}")

        return overlay

    # ------------------------------------------------------------------------------------------------------------------
    # ------------------------------------------------------------------------------------------------------------------
    def extract_masks(self, image_dir, output_dir):
        print("Extract masks from COCO file...")

        coco_annotation_file = os.path.join(image_dir, 'instances_default.json')
        coco = COCO(coco_annotation_file)

        os.makedirs(output_dir, exist_ok=True)

        all_results = []
        label_counter = Counter()
        coverage_stats = []

        # Build a distinct color palette sized to the number of categories
        cat_ids = coco.getCatIds()
        num_cats = len(cat_ids)
        palette = generate_colors(num_cats)
        color_map = {}

        img_ids = coco.getImgIds()

        # --- Ensure QApplication exists ---
        app = QApplication.instance()
        if app is None:
            app = QApplication(sys.argv)

        # --- Initialize progress wheel ---
        wheel = QProgressWheel(startVal=0, maxVal=len(img_ids), title="Analyzing Images", total=len(img_ids))

        # Flag to stop processing if wheel is closed
        self.stop_requested = False

        def on_wheel_closed():
            self.stop_requested = True

        wheel.destroyed.connect(lambda _: on_wheel_closed())

        for idx, img_id in enumerate(img_ids, start=1):
            if self.stop_requested:  # <-- break out if wheel closed
                print("Processing stopped by user.")
                break

            img_info = coco.loadImgs(img_id)[0]
            filename = img_info['file_name']
            base, ext = os.path.splitext(filename)
            img_path = os.path.join(image_dir, filename)

            image = cv2.imread(img_path)
            if image is None:
                print(f"Could not read image: {img_path}")
                wheel.setValue(idx)
                QApplication.processEvents()
                continue

            height, width, _ = image.shape
            gray_mask = np.zeros((height, width), dtype=np.uint8)
            color_mask = np.zeros((height, width, 3), dtype=np.uint8)

            ann_ids = coco.getAnnIds(imgIds=img_id)
            anns = coco.loadAnns(ann_ids)

            labels = []
            mask_area = 0

            for ann in anns:
                rle = coco.annToRLE(ann)
                binary_mask = maskUtils.decode(rle).astype(bool)

                gray_mask = np.maximum(gray_mask, binary_mask.astype(np.uint8) * 255)

                cat = coco.loadCats(ann['category_id'])[0]['name']
                labels.append(cat)
                label_counter[cat] += 1

                if cat not in color_map:
                    color_map[cat] = palette[len(color_map) % len(palette)]

                color_mask[binary_mask] = color_map[cat]
                mask_area += np.sum(binary_mask)

            # Save masks
            gray_path = os.path.join(output_dir, f"{base}_mask.png")
            colored_path = os.path.join(output_dir, f"{base}_color_mask.png")
            cv2.imwrite(gray_path, gray_mask)
            cv2.imwrite(colored_path, color_mask)

            # Save overlay
            overlay_img = self.create_overlay(image, color_mask, output_dir, base, ext=".png")

            # Collect coverage stats
            coverage_pct = (mask_area / (height * width)) * 100
            coverage_stats.append({
                "image_file": filename,
                "height": height,
                "width": width,
                "mask_area": mask_area,
                "coverage_pct": round(coverage_pct, 2),
                "num_labels": len(set(labels)),
                "labels": ", ".join(sorted(set(labels)))
            })

            all_results.append({
                "image_file": filename,
                "labels": ", ".join(sorted(set(labels)))
            })

            # --- Build analysis text for panel ---
            analysis_text = (
                f"Image: {filename}\n"
                f"Labels: {', '.join(sorted(set(labels)))}\n"
                f"Coverage: {coverage_pct:.2f}%\n"
                f"Resolution: {width}x{height}"
            )

            panel_path = os.path.join(output_dir, f"{base}_panel.png")
            create_panel(image, overlay_img, color_mask, analysis_text, panel_path, label_colors=color_map)

            # --- Update progress wheel ---
            wheel.setValue(idx)
            QApplication.processEvents()

        # --- Finish progress wheel ---
        if not self.stop_requested and wheel is not None:
            # update to 100% if still running
            wheel.setValue(len(img_ids))
            # don't force-close if the user already closed it
            try:
                wheel.close()
            except RuntimeError:
                # wheel was already closed by the user
                pass

        # Only save diagnostics if not stopped
        if not self.stop_requested:
            save_diagnostics(output_dir, label_counter, coverage_stats, all_results)
            report_analysis(output_dir, label_counter, coverage_stats, image_dir)


# ======================================================================================================================
# ======================================================================================================================
# ===   ===   ===   ===   ===   ===   ===   ===      HELPER FUNCTIONS      ===   ===   ===   ===   ===   ===   ===   ===
# ======================================================================================================================
# ======================================================================================================================
def save_diagnostics(output_dir, label_counter, coverage_stats, all_results):
    """Save label distribution, coverage stats, and summary Excel."""
    label_df = pd.DataFrame(label_counter.items(), columns=["label", "count"])
    label_df.to_csv(os.path.join(output_dir, "label_distribution.csv"), index=False)

    coverage_df = pd.DataFrame(coverage_stats)
    coverage_df.to_csv(os.path.join(output_dir, "mask_coverage_stats.csv"), index=False)

    summary_df = pd.DataFrame(all_results, columns=["image_file", "labels"])
    summary_df.to_excel(os.path.join(output_dir, "mask_labels.xlsx"), index=False)

# ----------------------------------------------------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------------------------------------
def analyze_label_distribution(label_counter):
    """Check label distribution for imbalance or missing labels."""
    issues = []
    if not label_counter:
        issues.append("No labels found in the dataset. Training will fail.")
    else:
        min_count = min(label_counter.values())
        max_count = max(label_counter.values())
        if min_count == 0:
            issues.append("Some expected labels never appear in the dataset.")
        if max_count > 10 * min_count:
            issues.append("Dataset is imbalanced. Some classes dominate while others are rare.")
    return issues

# ----------------------------------------------------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------------------------------------
def analyze_coverage(coverage_stats):
    """Check mask coverage for empty or low-coverage masks."""
    issues = []
    empty_masks = [row for row in coverage_stats if row["mask_area"] == 0]
    low_coverage = [row for row in coverage_stats if row["coverage_pct"] < 1.0]

    if empty_masks:
        issues.append(f"{len(empty_masks)} images have empty masks. These may indicate missing annotations.")
    if low_coverage:
        issues.append(f"{len(low_coverage)} images have masks covering less than 1 percent of the image.")
    return issues

# ----------------------------------------------------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------------------------------------
def analyze_dataset_size(num_images, min_threshold=50):
    issues = []
    if num_images < min_threshold:
        issues.append(
            f"Dataset has only {num_images} images. "
            f"Recommended minimum is {min_threshold} for segmentation tasks."
        )
    return issues

# ----------------------------------------------------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------------------------------------
def analyze_dataset_size(num_images, min_threshold=50):
    """Check if dataset has enough images."""
    issues = []
    if num_images < min_threshold:
        issues.append(
            f"Dataset has only {num_images} images. "
            f"Recommended minimum is {min_threshold} for segmentation tasks."
        )
    return issues

# ----------------------------------------------------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------------------------------------
def analyze_label_distribution(label_counter, imbalance_ratio=10, min_per_class=5):
    """Check label distribution for imbalance or missing labels."""
    issues = []
    if not label_counter:
        issues.append("No labels found in the dataset. Training will fail.")
    else:
        min_count = min(label_counter.values())
        max_count = max(label_counter.values())
        if min_count == 0:
            issues.append("Some expected labels never appear in the dataset.")
        if max_count > imbalance_ratio * min_count:
            issues.append("Dataset is imbalanced. Some classes dominate while others are rare.")
        for label, count in label_counter.items():
            if count < min_per_class:
                issues.append(f"Label '{label}' has only {count} examples, which may be too few.")
    return issues

# ----------------------------------------------------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------------------------------------
def analyze_coverage(coverage_stats, low_threshold=1.0):
    """Check mask coverage for empty or low-coverage masks."""
    issues = []
    empty_masks = [row for row in coverage_stats if row["mask_area"] == 0]
    low_coverage = [row for row in coverage_stats if row["coverage_pct"] < low_threshold]

    if empty_masks:
        issues.append(f"{len(empty_masks)} images have empty masks. These may indicate missing annotations.")
    if low_coverage:
        issues.append(f"{len(low_coverage)} images have masks covering less than {low_threshold} percent of the image.")
    return issues

# ----------------------------------------------------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------------------------------------
def analyze_annotation_density(coverage_stats, min_annotations=1):
    """Check average number of labels per image."""
    issues = []
    low_density = [row for row in coverage_stats if row["num_labels"] < min_annotations]
    if low_density:
        issues.append(f"{len(low_density)} images have fewer than {min_annotations} labels.")
    return issues

# ----------------------------------------------------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------------------------------------
def analyze_resolution_consistency(coverage_stats, tolerance=0.5):
    """Check if image resolutions are consistent."""
    issues = []
    if not coverage_stats:
        return issues

    heights = [row["height"] for row in coverage_stats]
    widths = [row["width"] for row in coverage_stats]

    avg_h, avg_w = np.mean(heights), np.mean(widths)
    for row in coverage_stats:
        if abs(row["height"] - avg_h) / avg_h > tolerance or abs(row["width"] - avg_w) / avg_w > tolerance:
            issues.append(
                f"Image {row['image_file']} has resolution {row['width']}x{row['height']} "
                f"which differs significantly from average {int(avg_w)}x{int(avg_h)}."
            )
    return issues

# ----------------------------------------------------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------------------------------------
def analyze_mask_alignment(image_dir, coverage_stats):
    """Check if masks exist and match image dimensions."""
    issues = []
    for row in coverage_stats:
        mask_path = os.path.join(image_dir, f"{os.path.splitext(row['image_file'])[0]}_mask.png")
        if not os.path.exists(mask_path):
            issues.append(f"Mask file missing for image {row['image_file']}.")
        else:
            mask = cv2.imread(mask_path, cv2.IMREAD_GRAYSCALE)
            if mask is None:
                issues.append(f"Mask file for {row['image_file']} could not be read.")
            elif mask.shape != (row["height"], row["width"]):
                issues.append(
                    f"Mask for {row['image_file']} has shape {mask.shape}, "
                    f"but image is {row['width']}x{row['height']}."
                )
    return issues

# ----------------------------------------------------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------------------------------------
def analyze_file_integrity(image_dir, coverage_stats):
    """Check if all images can be read properly."""
    issues = []
    for row in coverage_stats:
        img_path = os.path.join(image_dir, row["image_file"])
        image = cv2.imread(img_path)
        if image is None:
            issues.append(f"Image {row['image_file']} could not be read from disk.")
    return issues

# ----------------------------------------------------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------------------------------------
def report_analysis(output_dir, label_counter, coverage_stats, image_dir):
    report_lines = []
    report_lines.append("\nDataset Analysis Report")
    issues = []
    issues.extend(analyze_dataset_size(len(coverage_stats)))
    issues.extend(analyze_label_distribution(label_counter))
    issues.extend(analyze_coverage(coverage_stats))
    issues.extend(analyze_annotation_density(coverage_stats))
    issues.extend(analyze_resolution_consistency(coverage_stats))
    issues.extend(analyze_mask_alignment(output_dir, coverage_stats))
    issues.extend(analyze_file_integrity(image_dir, coverage_stats))

    if issues:
        for issue in issues:
            print(issue)
            report_lines.append(issue)
    else:
        success_msg = (
            "No major issues detected. Dataset looks good for training.\n"
            "Tests performed:\n"
            "- Dataset size check\n"
            "- Label distribution check\n"
            "- Mask coverage check\n"
            "- Annotation density check\n"
            "- Resolution consistency check\n"
            "- Mask alignment check\n"
            "- File integrity check\n"
        )
        print(success_msg)
        report_lines.append(success_msg)

    final_msg = f"\nSaved masks, overlays, label distribution, and diagnostics to: {output_dir}"
    print(final_msg)
    report_lines.append(final_msg)

    # --- Write to text file ---
    analysis_path = os.path.join(output_dir, "training_data_analysis.txt")
    with open(analysis_path, "w", encoding="utf-8") as f:
        for line in report_lines:
            f.write(line + "\n")

# ----------------------------------------------------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------------------------------------
def create_panel(original_img, overlay_img, color_mask, analysis_text, output_path, label_colors=None):
    """
    Create a 2x2 panel with:
    [Original | Overlay]
    [Color Mask | Analysis Text with color swatches]

    Parameters
    ----------
    original_img : np.ndarray
        Original image (BGR).
    overlay_img : np.ndarray
        Overlay image (BGR).
    color_mask : np.ndarray
        Color mask image (BGR).
    analysis_text : str
        Text summary of analysis results.
    output_path : str
        Path to save the final panel image.
    label_colors : dict[str, tuple[int,int,int]], optional
        Mapping {label: (B,G,R)} used for swatches.
    """

    # --- Ensure all images are the same size ---
    height, width = original_img.shape[:2]

    def resize(img):
        return cv2.resize(img, (width, height))

    overlay_img = resize(overlay_img)
    color_mask = resize(color_mask)

    # --- Create text panel ---
    text_panel = np.ones((height, width, 3), dtype=np.uint8) * 255  # white background
    y0, dy = 30, 30

    for i, line in enumerate(analysis_text.split("\n")):
        y = y0 + i * dy

        # If this line starts with "Labels:" and we have label_colors, draw swatches
        if line.startswith("Labels:") and label_colors:
            # Extract labels from line
            labels = [lbl.strip() for lbl in line.replace("Labels:", "").split(",") if lbl.strip()]
            x_text = 20
            cv2.putText(text_panel, "Labels:", (x_text, y),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 0, 0), 2, cv2.LINE_AA)
            x_cursor = x_text + 120
            for lbl in labels:
                color = label_colors.get(lbl, (0, 0, 0))
                # Draw swatch rectangle
                cv2.rectangle(text_panel, (x_cursor, y-20), (x_cursor+40, y+5), color, -1)
                # Draw label text next to swatch
                cv2.putText(text_panel, lbl, (x_cursor+50, y),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 0, 0), 2, cv2.LINE_AA)
                x_cursor += 150
        else:
            cv2.putText(text_panel, line, (20, y),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 0, 0), 2, cv2.LINE_AA)

    # --- Stack into 2x2 grid ---
    top_row = np.hstack((original_img, overlay_img))
    bottom_row = np.hstack((color_mask, text_panel))
    panel = np.vstack((top_row, bottom_row))

    # --- Save result ---
    cv2.imwrite(output_path, panel)
    print(f"Panel saved: {output_path}")

# ----------------------------------------------------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------------------------------------
def generate_colors(n):
    colors = []
    for i in range(n):
        hue = i / n
        r, g, b = [int(c*255) for c in plt.cm.hsv(hue)[:3]]
        colors.append((b, g, r))  # convert to BGR for OpenCV
    return colors
