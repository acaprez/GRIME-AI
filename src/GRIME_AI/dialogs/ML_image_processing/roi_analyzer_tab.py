#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# ROIAnalyzerTab.py
# Direct port of the ROI Analyzer functionality and signal wiring from GRIME_AI_ML_ImageProcessingDlg
# Author: John Edward Stranzl, Jr.
# License: Apache License, Version 2.0

import os
import re
import cv2
import pandas as pd
import seaborn as sns
import matplotlib.pyplot as plt
import statsmodels.api as sm

from pathlib import Path
from PyQt5 import QtCore
from PyQt5.QtCore import Qt, QTimer
from PyQt5.QtWidgets import QWidget, QFileDialog, QListWidgetItem, QMessageBox
from PyQt5.QtGui import QPixmap, QIcon, QImage, QPainter, QColor

from GRIME_AI.GRIME_AI_JSON_Editor import JsonEditor
from GRIME_AI.GRIME_AI_QProgressWheel import QProgressWheel
from GRIME_AI.GRIME_AI_CSS_Styles import BUTTON_CSS_STEEL_BLUE


class ROIAnalyzerTab(QWidget):

    def __init__(self, parent=None):
        super().__init__(parent)

        # EXPECTED UI WIDGETS (assign before calling wire_connections()):
        # - self.lineEdit_ROI_images_folder
        # - self.pushButton_browse_ROI_images_folder
        # - self.pushButton_analyze
        # - self.pushButton_extract_ROI_features
        # - self.listWidget_filmstrip
        # - self.spinBox_numClusters
        # - self.label_displayImages
        # - self.lineEdit_intensity
        # - self.lineEdit_entropy
        # - self.lineEdit_Texture
        # - self.lineEdit_GLI
        # - self.lineEdit_GCC
        # - self.buttonBox_close  (optional; used to close in original dialog)

        # Runtime state used by ROI Analyzer
        self._pairs = []
        self._pendingThumbnails = []
        self._batchSize = 10
        self._batchDelay = 50
        self._loadToken = 0
        self.num_clusters = None

        # Dataframes created during feature extraction and alignment
        self.roi_metrics_df = None
        self.related_csv_df = None
        self.aligned_df = None

        # Preserve full-size canvas for resize scaling
        self._full_canvas = None

    # ------------------------------------------------------------------------------------------------------------------
    # ------------------------------------------------------------------------------------------------------------------
    def configure_filmstrip(self):
        """
        Force the filmstrip to be exactly one thumbnail high,
        no wrapping, no extra margins or space.
        """
        lw = self.listWidget_filmstrip

        # No wrapping, no spacing
        lw.setWrapping(False)
        lw.setSpacing(0)

        # Remove margins
        lw.setContentsMargins(0, 0, 0, 0)
        lw.setViewportMargins(0, 0, 0, 0)

        # Remove frame
        from PyQt5.QtWidgets import QFrame
        lw.setFrameShape(QFrame.NoFrame)

        # Fix the height to exactly one icon row
        icon_h = lw.iconSize().height()
        lw.setFixedHeight(icon_h)

    # ------------------------------------------------------------------------------------------------------------------
    # Signal wiring
    # ------------------------------------------------------------------------------------------------------------------
    def wire_connections(self):
        # Folder text field
        self.lineEdit_ROI_images_folder.editingFinished.connect(self._on_roi_images_folder_changed)

        # Browse button
        self.pushButton_browse_ROI_images_folder.clicked.connect(self.browse_ROI_images_folder)
        self.pushButton_browse_ROI_images_folder.setStyleSheet(BUTTON_CSS_STEEL_BLUE)

        # Analyze button
        self.pushButton_analyze.clicked.connect(self.analyze_roi)
        self.pushButton_analyze.setStyleSheet(BUTTON_CSS_STEEL_BLUE)

        # Extract features button
        self.pushButton_extract_ROI_features.clicked.connect(self.extract_ROI_features)
        self.pushButton_extract_ROI_features.setStyleSheet(BUTTON_CSS_STEEL_BLUE)

        # Filmstrip click
        self.listWidget_filmstrip.itemClicked.connect(self.on_filmstrip_item_clicked)

        # Cluster count spin box
        self.num_clusters = self.spinBox_numClusters.value()
        self.spinBox_numClusters.valueChanged.connect(self._on_num_clusters_changed)

        # Optional close hook (matches original)
        if hasattr(self, "buttonBox_close") and self.buttonBox_close is not None:
            self.buttonBox_close.rejected.connect(self.reject)

        # Populate folder from config (matches original behavior)
        roi_analyzer_image_folder = JsonEditor().getValue("ROI_Analyzer_Images_Folder")
        if roi_analyzer_image_folder:
            self.lineEdit_ROI_images_folder.setText(roi_analyzer_image_folder)

    # ******************************************************************************************************************
    # *   ROI ANALYZER LOGIC                                                                    *
    # ******************************************************************************************************************
    # ------------------------------------------------------------------------------------------------------------------
    # ------------------------------------------------------------------------------------------------------------------
    def _on_num_clusters_changed(self, value):
        self.num_clusters = value

        current = self.listWidget_filmstrip.currentItem()
        if current:
            # re-run analysis on the highlighted image
            self.on_filmstrip_item_clicked(current)
        else:
            # fallback: rerun the initial analysis sequence
            self.analyze_roi()

    # ------------------------------------------------------------------------------------------------------------------
    # ------------------------------------------------------------------------------------------------------------------
    def on_filmstrip_item_clicked(self, item: QListWidgetItem):
        """
        When a thumbnail is clicked, re-run ROI analysis on that image/mask pair
        and update label_displayImages with the composite plot + top-3 swatches.
        """
        # Retrieve the index we stored in populate_filmstrip
        idx = item.data(Qt.UserRole)
        orig_path, mask_path = self._pairs[idx]

        # grab the user-selected cluster count
        n_clusters = self.spinBox_numClusters.value()

        # Run analysis for this specific pair
        from GRIME_AI.GRIME_AI_ROI_Analyzer import GRIME_AI_ROI_Analyzer
        analyzer = GRIME_AI_ROI_Analyzer(orig_path, mask_path, clusters=n_clusters)
        analyzer.run_analysis()

        # ─── populate metric fields ───
        self.lineEdit_intensity.setText(f"{analyzer.roi_intensity:.2f}")
        self.lineEdit_entropy.setText(f"{analyzer.roi_entropy:.4f}")
        self.lineEdit_Texture.setText(f"{analyzer.roi_texture:.4f}")
        self.lineEdit_GLI.setText(f"{analyzer.mean_gli:.2f}")
        self.lineEdit_GCC.setText(f"{analyzer.mean_gcc:.2f}")

        # Display composite+metrics plot
        composite_pix = analyzer.get_results_pixmap()
        self.label_displayImages.setPixmap(composite_pix)

        # Overlay top-3 dominant color swatches
        n_clusters = self.spinBox_numClusters.value()
        swatches = analyzer.dominant_rgb_list[:n_clusters]
        self._draw_color_swatches_on_label(swatches, swatch_size=100)

    # ------------------------------------------------------------------------------------------------------------------
    # ------------------------------------------------------------------------------------------------------------------
    def populate_filmstrip(self, image_paths):
        """
        1) Clear old thumbnails.
        2) Enqueue placeholders + paths.
        3) Kick off _loadNextBatch via QTimer.
        """
        lw = self.listWidget_filmstrip
        lw.clear()

        # Invalidate any previous loader
        self._loadToken += 1
        token = self._loadToken
        self._pendingThumbnails.clear()

        # Create one blank item per image path
        iconSize = lw.iconSize()
        for idx, path in enumerate(image_paths):
            item = QListWidgetItem(QIcon(), "")
            item.setData(Qt.UserRole, idx)
            item.setSizeHint(iconSize)
            lw.addItem(item)
            self._pendingThumbnails.append((item, path, token))

        # Highlight first by default
        if lw.count():
            lw.setCurrentRow(0)

        # Schedule the first batch load
        QTimer.singleShot(self._batchDelay, lambda: self._loadNextBatch(token))

    # ------------------------------------------------------------------------------------------------------------------
    # ------------------------------------------------------------------------------------------------------------------
    def _loadNextBatch(self, token):
        """
        Pop up to self._batchSize thumbnails from self._pendingThumbnails,
        assign icons, then reschedule if more remain.
        """
        # Cancel if stale
        if token != self._loadToken:
            return

        lw = self.listWidget_filmstrip
        iconSize = lw.iconSize()

        for _ in range(min(self._batchSize, len(self._pendingThumbnails))):
            item, path, _ = self._pendingThumbnails.pop(0)
            if not os.path.exists(path):
                continue
            pix = QPixmap(path)
            if pix.isNull():
                continue

            thumb = pix.scaled(
                iconSize,
                Qt.KeepAspectRatio,
                Qt.SmoothTransformation
            )
            item.setIcon(QIcon(thumb))

        # More to do?
        if self._pendingThumbnails:
            QTimer.singleShot(self._batchDelay, lambda: self._loadNextBatch(token))

    # ------------------------------------------------------------------------------------------------------------------
    # ------------------------------------------------------------------------------------------------------------------
    def browse_ROI_images_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "Select Folder", os.getcwd())
        if not folder:
            return
        self.lineEdit_ROI_images_folder.setText(folder)

        self._on_roi_images_folder_changed()

    # ------------------------------------------------------------------------------------------------------------------
    # ------------------------------------------------------------------------------------------------------------------
    def _on_roi_images_folder_changed(self):
        folder = self.lineEdit_ROI_images_folder.text().strip()
        if not folder or not os.path.isdir(folder):
            return

        JsonEditor().update_json_entry("ROI_Analyzer_Images_Folder", folder)

        from GRIME_AI.GRIME_AI_ROI_Analyzer import GRIME_AI_ROI_Analyzer
        temp = GRIME_AI_ROI_Analyzer("", "")
        pairs = temp.generate_file_pairs(folder)
        if not pairs:
            return

        self._pairs = pairs
        self.populate_filmstrip([orig for orig, _ in pairs])

    # ------------------------------------------------------------------------------------------------------------------
    # ------------------------------------------------------------------------------------------------------------------
    def analyze_roi(self):
        """
        1) Generate file pairs and populate filmstrip
        2) Run analysis on the first (or clicked) pair
        3) Display composite plot, fallback to original if empty
        4) Overlay top-3 color swatches
        """
        folder = self.lineEdit_ROI_images_folder.text().strip()
        if not folder:
            QMessageBox.warning(self, "ROI Analyzer", "Please specify a folder path.")
            return

        try:
            from GRIME_AI.GRIME_AI_ROI_Analyzer import GRIME_AI_ROI_Analyzer
        except ImportError:
            QMessageBox.warning(self, "ROI Analyzer", "Unable to import ROI Analyzer module.")
            return

        # 1) generate pairs + batched filmstrip population
        temp = GRIME_AI_ROI_Analyzer("", "")
        pairs = temp.generate_file_pairs(folder)
        if not pairs:
            QMessageBox.warning(self, "ROI Analyzer", "No image/mask pairs found.")
            return
        self._pairs = pairs

        # Replace inline loop with batched loader
        image_paths = [orig for orig, _ in pairs]
        self.populate_filmstrip(image_paths)

        # 2) analyze the first pair (index 0) by default
        orig_path, mask_path = pairs[0]
        n_clusters = self.spinBox_numClusters.value()
        analyzer = GRIME_AI_ROI_Analyzer(orig_path, mask_path, clusters=n_clusters)
        analyzer.run_analysis()

        # 3) try to get the composite+metrics pixmap
        try:
            comp_pix = analyzer.get_results_pixmap()
        except Exception as e:
            # fallback: load the original image manually
            img = cv2.imread(orig_path)
            if img is None or img.size == 0:
                QMessageBox.warning(
                    self,
                    "ROI Analyzer",
                    f"Could not generate results pixmap or load original:\n{e}"
                )
                return

            rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
            h, w = rgb.shape[:2]
            comp_pix = QPixmap.fromImage(
                QImage(rgb.data, w, h, w * 3, QImage.Format_RGB888)
            )

        self.label_displayImages.setPixmap(comp_pix)

        # ─── populate metric fields ───
        self.lineEdit_intensity.setText(f"{analyzer.roi_intensity:.2f}")
        self.lineEdit_entropy.setText(f"{analyzer.roi_entropy:.4f}")
        self.lineEdit_Texture.setText(f"{analyzer.roi_texture:.4f}")
        self.lineEdit_GLI.setText(f"{analyzer.mean_gli:.2f}")
        self.lineEdit_GCC.setText(f"{analyzer.mean_gcc:.2f}")

        # 4) overlay the top-3 swatches
        n_clusters = self.spinBox_numClusters.value()
        swatches = getattr(analyzer, "dominant_rgb_list", [])[:n_clusters]
        self._draw_color_swatches_on_label(swatches, swatch_size=100)

    # ------------------------------------------------------------------------------------------------------------------
    # ------------------------------------------------------------------------------------------------------------------
    def analyze_responses(self, df, predictors, responses, output_folder):
        """
        Analyze correlations and regressions between predictors and responses,
        with numeric coercion and plot skipping for insufficient data.
        Saves plots and correlation tables to the specified output folder.
        """
        output_dir = Path(output_folder)
        output_dir.mkdir(parents=True, exist_ok=True)

        # --- 1. Coerce predictors and responses to numeric ---
        df1 = df.copy()
        df1[predictors + responses] = df1[predictors + responses].apply(pd.to_numeric, errors='coerce')

        # --- 2. Distribution plots ---
        for col in predictors + responses:
            plt.figure()
            sns.histplot(df[col].dropna(), kde=True)
            plt.title(f"Distribution of {col}")
            plt.tight_layout()
            plt.savefig(output_dir / f"hist_{col}.png")
            plt.close()

        # --- 3. Correlations ---
        pearson_results = {}
        spearman_results = {}
        for resp in responses:
            pearson_results[resp] = df1[predictors + [resp]].corr(method='pearson')[resp].drop(resp)
            spearman_results[resp] = df1[predictors + [resp]].corr(method='spearman')[resp].drop(resp)

        pearson_df = pd.DataFrame(pearson_results)
        spearman_df = pd.DataFrame(spearman_results)

        # --- 4. Scatterplots with regression lines ---
        for resp in responses:
            for pred in predictors:
                x = df1[pred]
                y = df1[resp]
                valid = x.notna() & y.notna()
                if valid.sum() < 2:
                    print(f"Skipping plot for {pred} vs {resp}: insufficient numeric data")
                    continue

                plt.figure()
                sns.regplot(x=x[valid], y=y[valid], scatter_kws={'alpha': 0.5})
                plt.xlabel(pred)
                plt.ylabel(resp)
                plt.title(f"{pred} vs {resp}")
                plt.tight_layout()
                plt.savefig(output_dir / f"scatter_{pred}_vs_{resp}.png")
                plt.close()

        # --- 5. Multiple linear regression models ---
        model_summaries = {}
        for resp in responses:
            X = df1[predictors]
            y = df1[resp]
            valid = X.notna().all(axis=1) & y.notna()
            if valid.sum() < 2:
                print(f"Skipping regression for {resp}: insufficient numeric data")
                continue
            X_valid = sm.add_constant(X[valid])
            y_valid = y[valid]
            model = sm.OLS(y_valid, X_valid).fit()
            model_summaries[resp] = model.summary().as_text()
            print(f"\n=== Regression for {resp} ===")
            print(model.summary())

        # --- 6. Correlation heatmap ---
        plt.figure(figsize=(10, 8))
        corr_matrix = df1[predictors + responses].corr(method='pearson')
        sns.heatmap(corr_matrix, annot=True, fmt=".2f", cmap="coolwarm")
        plt.title("Correlation Heatmap (Pearson)")
        plt.tight_layout()
        plt.savefig(output_dir / "correlation_heatmap.png")
        plt.close()

        # --- 7. Save correlation tables and model summaries to Excel ---
        excel_path = output_dir / "analysis_results.xlsx"
        with pd.ExcelWriter(excel_path) as writer:
            pearson_df.to_excel(writer, sheet_name="Pearson_Corr")
            spearman_df.to_excel(writer, sheet_name="Spearman_Corr")
            summary_df = pd.DataFrame.from_dict(model_summaries, orient='index', columns=['Summary'])
            summary_df.to_excel(writer, sheet_name="Model_Summaries")

        print(f"\nAnalysis complete. Results saved to {output_dir}")

    # ------------------------------------------------------------------------------------------------------------------
    # ------------------------------------------------------------------------------------------------------------------
    def extract_ROI_features(self):
        # GET SETTINGS FROM UI CONTROLS
        n_clusters = self.spinBox_numClusters.value()

        # Regex pattern for datetime stamp in filenames, e.g., 2025-09-20T16-30-02Z
        dt_pattern = re.compile(r"(\d{4}-\d{2}-\d{2})T(\d{2}-\d{2}-\d{2})Z")

        # 1) Get and validate output folder path
        output_folder = self.lineEdit_ROI_images_folder.text().strip()
        if not output_folder:
            QMessageBox.warning(self, "No Output Folder", "Please specify an output folder.")
            return

        os.makedirs(output_folder, exist_ok=True)

        # Ensure we have image/mask pairs
        if not getattr(self, "_pairs", None):
            QMessageBox.warning(self, "No Image Pairs", "No (image, mask) pairs found.")
            return

        # Resolve first image path and image_folder early
        first_img_path = self._pairs[0][0]
        image_folder = os.path.dirname(first_img_path)

        # Determine a datetime prefix for filenames from the FIRST image that matches the pattern
        file_dt_prefix = "no_datetime"
        for img_path, _ in self._pairs:
            img_name = os.path.basename(img_path)
            match_prefix_dt = dt_pattern.search(img_name)
            if match_prefix_dt:
                # Example output: 2025-09-20_163002
                file_dt_prefix = f"{match_prefix_dt.group(1)}_{match_prefix_dt.group(2).replace('-', '')}"
                break

        # Output file paths with datetime prefix
        csv_path = os.path.join(output_folder, f"{file_dt_prefix}_roi_metrics.csv")
        xlsx_path = os.path.join(output_folder, f"{file_dt_prefix}_roi_metrics.xlsx")
        aligned_xlsx_path = os.path.join(output_folder, f"{file_dt_prefix}_aligned_results.xlsx")

        # 2) Prepare header and container for all rows
        header = [
            "Image Path",
            "Mask Path",
            "Capture Date",
            "Capture Time",
            "ROI Intensity",
            "ROI Entropy",
            "ROI Texture",
            "Mean GLI",
            "Mean GCC",
            "ROI Pixel Count",
            "ROI Area",
            "Image Height",
            "Image Width",
            "Image Total Pixels",
            "ROI Area Percentage",
        ]

        # Add HSV columns for each cluster
        for i in range(1, n_clusters + 1):
            header.extend([
                f"Cluster {i} H",
                f"Cluster {i} S",
                f"Cluster {i} V",
            ])

        # (Optional) Add RGB columns for each cluster
        for i in range(1, n_clusters + 1):
            header.extend([
                f"Cluster {i} R",
                f"Cluster {i} G",
                f"Cluster {i} B",
            ])

        rows = [header]

        # --- Find related CSV file (CSV has " - ", JPGs don't). Prefix is JPG stem before the datetime stamp. ---
        first_image_stem = os.path.splitext(os.path.basename(first_img_path))[0]
        match_dt = dt_pattern.search(first_image_stem)
        if match_dt:
            common_prefix = first_image_stem[:match_dt.start()].rstrip("_ ")
        else:
            common_prefix = first_image_stem

        matching_csv_path = None
        try:
            for fname in os.listdir(image_folder):
                # Case-insensitive check; CSV must start with "<common_prefix> - "
                if fname.lower().endswith(".csv") and fname.lower().startswith((common_prefix + " - ").lower()):
                    matching_csv_path = os.path.join(image_folder, fname)
                    break
        except Exception as e:
            print(f"Error listing directory {image_folder}: {e}")
            matching_csv_path = None

        if matching_csv_path:
            try:
                self.related_csv_df = pd.read_csv(matching_csv_path)

                # --- Split 'datetime' column into Date and Time (24-hour) ---
                if 'datetime' in self.related_csv_df.columns:
                    self.related_csv_df['datetime'] = pd.to_datetime(
                        self.related_csv_df['datetime'],
                        format="%m/%d/%Y %H:%M",
                        errors="coerce"
                    )
                    self.related_csv_df['Date'] = self.related_csv_df['datetime'].dt.strftime("%Y-%m-%d")
                    self.related_csv_df['Time'] = self.related_csv_df['datetime'].dt.strftime("%H:%M:%S")

            except Exception as e:
                print(f"Error reading related CSV {matching_csv_path}: {e}")
                self.related_csv_df = None
        else:
            self.related_csv_df = None

        progress_bar_closed = False
        total_iterations = len(self._pairs)
        progressBar = QProgressWheel(
            title="Extracting features in-progress...",
            total=total_iterations,
            on_close=lambda: setattr(self, "progress_bar_closed", True)
        )

        # 3) Iterate pairs, run analysis, collect results
        try:
            for i, (orig_path, mask_path) in enumerate(self._pairs):
                progressBar.setValue(i)

                # Extract datetime from image filename
                filename = os.path.basename(orig_path)
                match = dt_pattern.search(filename)
                if match:
                    capture_date = match.group(1)
                    capture_time = match.group(2).replace("-", ":")
                else:
                    capture_date = ""
                    capture_time = ""

                from GRIME_AI.GRIME_AI_ROI_Analyzer import GRIME_AI_ROI_Analyzer
                analyzer = GRIME_AI_ROI_Analyzer(orig_path, mask_path, clusters=n_clusters)

                try:
                    analyzer.run_analysis()
                except Exception as e:
                    print(f"Failed on {orig_path}, {mask_path}: {e}")
                    continue

                # Build row: fixed ROI metrics first
                data_row = [
                    orig_path,
                    mask_path,
                    capture_date,
                    capture_time,
                    f"{analyzer.roi_intensity:.2f}",
                    f"{analyzer.roi_entropy:.4f}",
                    f"{analyzer.roi_texture:.4f}",
                    f"{analyzer.mean_gli:.2f}",
                    f"{analyzer.mean_gcc:.2f}",
                    f"{analyzer.ROI_total_pixels:.2f}",
                    f"{analyzer.ROI_total_area:.2f}",
                    f"{analyzer.image_height:.2f}",
                    f"{analyzer.image_width:.2f}",
                    f"{analyzer.image_total_pixels:.2f}",
                    f"{analyzer.ROI_percentage:.2f}",
                ]

                # Append HSV values for each cluster
                for (h, s, v) in analyzer.dominant_hsv_list:
                    data_row.extend([f"{h:.4f}", f"{s:.4f}", f"{v:.4f}"])

                # Append RGB values for each cluster
                for (r, g, b) in analyzer.dominant_rgb_list:
                    data_row.extend([f"{r:.4f}", f"{g:.4f}", f"{b:.4f}"])

                rows.append(data_row)

        finally:
            try:
                progressBar.close()
            except Exception:
                pass
            del progressBar

        # 3.5) Convert to DataFrame for in-memory use
        df = pd.DataFrame(rows[1:], columns=header)
        self.roi_metrics_df = df  # Keep in memory for subsequent steps

        # Print first 5 rows of each DataFrame
        print("\nFirst 5 rows of ROI Metrics DataFrame:")
        print(self.roi_metrics_df.head())
        if self.related_csv_df is not None:
            print("\nFirst 5 rows of Related CSV DataFrame:")
            print(self.related_csv_df.head())
        else:
            print("\nNo related CSV DataFrame to display.")

        # 4) Write out CSV of ROI metrics
        try:
            df.to_csv(csv_path, index=False)
        except Exception as e:
            print(f"Error writing CSV to {csv_path}: {e}")

        # 5) Write out XLSX with hyperlinks for ROI metrics
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font

            wb = Workbook()
            ws = wb.active
            ws.title = "ROI Metrics"

            # Write header row
            for col_idx, title in enumerate(header, start=1):
                ws.cell(row=1, column=col_idx, value=title)

            hyperlink_style = Font(color="0000FF", underline="single")

            for row_idx, data in enumerate(rows[1:], start=2):
                (
                    orig_full, mask_full, capture_date, capture_time,
                    intensity, entropy, texture, gli, gcc,
                    pixel_count, pixel_area, image_height,
                    image_width, image_total_pixels, roi_area_percentage,
                    *cluster_values
                ) = data

                img_name = os.path.basename(orig_full)
                cell_img = ws.cell(row=row_idx, column=1, value=img_name)
                cell_img.hyperlink = orig_full
                cell_img.font = hyperlink_style

                mask_name = os.path.basename(mask_full)
                cell_mask = ws.cell(row=row_idx, column=2, value=mask_name)
                cell_mask.hyperlink = mask_full
                cell_mask.font = hyperlink_style

                ws.cell(row=row_idx, column=3, value=capture_date)
                ws.cell(row=row_idx, column=4, value=capture_time)
                ws.cell(row=row_idx, column=5, value=float(intensity))
                ws.cell(row=row_idx, column=6, value=float(entropy))
                ws.cell(row=row_idx, column=7, value=float(texture))
                ws.cell(row=row_idx, column=8, value=float(gli))
                ws.cell(row=row_idx, column=9, value=float(gcc))
                ws.cell(row=row_idx, column=10, value=float(pixel_count))
                ws.cell(row=row_idx, column=11, value=float(pixel_area))
                ws.cell(row=row_idx, column=12, value=float(image_height))
                ws.cell(row=row_idx, column=13, value=float(image_width))
                ws.cell(row=row_idx, column=14, value=float(image_total_pixels))
                ws.cell(row=row_idx, column=15, value=float(roi_area_percentage))

                for offset, value in enumerate(cluster_values, start=16):
                    ws.cell(row=row_idx, column=offset, value=float(value))

            wb.save(xlsx_path)

        except ImportError:
            QMessageBox.warning(
                self,
                "XLSX Export Skipped",
                "The 'openpyxl' library is not installed. Install it via 'pip install openpyxl' to enable XLSX export."
            )
        except Exception as e:
            print(f"Error writing XLSX to {xlsx_path}: {e}")

        # 6) Align ROI metrics with related CSV by nearest timestamp and export aligned results
        aligned_written = False
        aligned_df = []

        try:
            if self.related_csv_df is not None and not self.related_csv_df.empty:
                roi_dt = pd.to_datetime(
                    self.roi_metrics_df["Capture Date"].astype(str).str.strip() + " " +
                    self.roi_metrics_df["Capture Time"].astype(str).str.strip(),
                    errors="coerce",
                    format="%Y-%m-%d %H:%M:%S"
                )
                roi_df = self.roi_metrics_df.copy()
                roi_df["datetime"] = roi_dt

                csv_df = self.related_csv_df.copy()
                if {"Date", "Time"}.issubset(csv_df.columns):
                    csv_dt = pd.to_datetime(
                        csv_df["Date"].astype(str).str.strip() + " " +
                        csv_df["Time"].astype(str).str.strip(),
                        errors="coerce",
                        format="%Y-%m-%d %H:%M:%S"
                    )
                    csv_df["datetime"] = csv_dt
                elif "datetime" in csv_df.columns:
                    csv_df["datetime"] = pd.to_datetime(csv_df["datetime"], errors="coerce")
                else:
                    csv_df["datetime"] = pd.NaT

                roi_df = roi_df.dropna(subset=["datetime"]).sort_values("datetime")
                csv_df = csv_df.dropna(subset=["datetime"]).sort_values("datetime")

                if not roi_df.empty and not csv_df.empty:
                    aligned_df = pd.merge_asof(
                        roi_df,
                        csv_df,
                        on="datetime",
                        direction="nearest",
                        tolerance=pd.Timedelta("6H")
                    )
                    aligned_df["Image Path"] = aligned_df["Image Path"].apply(os.path.basename)
                    aligned_df["Mask Path"] = aligned_df["Mask Path"].apply(os.path.basename)

                    #self.aligned_df = aligned_df

                    aligned_df.to_excel(aligned_xlsx_path, index=False)
                    aligned_written = True
                else:
                    print("Alignment skipped: one of the DataFrames has no valid datetime rows.")
            else:
                print("Alignment skipped: related_csv_df is missing or empty.")
        except Exception as e:
            print(f"Error during alignment or writing aligned XLSX: {e}")

        predictors = []
        for i in range(1, n_clusters + 1):
            predictors.extend([f'Cluster {i} H', f'Cluster {i} S', f'Cluster {i} V'])

        responses = ['Gage Height', 'Discharge']

        if aligned_df:
            self.analyze_responses(aligned_df, predictors=predictors, responses=responses, output_folder=output_folder)

        from os.path import exists as _exists
        QMessageBox.information(
            self,
            "Export Complete",
            f"Metrics written to:\n{csv_path if _exists(csv_path) else '(CSV write failed)'}\n"
            + (f"{xlsx_path}" if _exists(xlsx_path) else "(XLSX skipped or write failed)")
            + ("\nAligned results written to:\n" + aligned_xlsx_path if aligned_written and _exists(aligned_xlsx_path)
               else "\nAligned results skipped or write failed")
            + ("\nRelated CSV loaded into self.related_csv_df" if self.related_csv_df is not None else "\nNo related CSV found")
        )

    # ------------------------------------------------------------------------------------------------------------------
    # Display helpers (ported verbatim)
    # ------------------------------------------------------------------------------------------------------------------
    def display_color_swatches(self, rgb_list, swatch_size=100):
        count = len(rgb_list)
        if count == 0:
            return

        pixmap = QPixmap(swatch_size * count, swatch_size)
        pixmap.fill(Qt.transparent)

        painter = QPainter(pixmap)
        for i, (r, g, b) in enumerate(rgb_list):
            color = QColor(r, g, b)
            painter.fillRect(i * swatch_size, 0, swatch_size, swatch_size, color)
        painter.end()

        self.label_displayImages.setPixmap(pixmap)

    def _draw_color_swatches_on_label(self, rgb_list, swatch_size=100):
        base = self.label_displayImages.pixmap()
        if base:
            base_w, base_h = base.width(), base.height()
        else:
            base_w, base_h = 0, 0

        swatch_count = len(rgb_list)
        swatches_w = swatch_count * swatch_size
        total_w = max(base_w, swatches_w)
        total_h = base_h + (swatch_size if swatch_count else 0)

        canvas = QPixmap(total_w, total_h)
        canvas.fill(Qt.transparent)

        painter = QPainter(canvas)
        if base:
            painter.drawPixmap(0, 0, base)

        for i, (r, g, b) in enumerate(rgb_list):
            x = i * swatch_size
            y = base_h
            painter.fillRect(x, y, swatch_size, swatch_size, QColor(r, g, b))
        painter.end()

        self._full_canvas = canvas

        scaled = self._full_canvas.scaled(
            self.label_displayImages.size(),
            Qt.KeepAspectRatio,
            Qt.SmoothTransformation
        )
        self.label_displayImages.setPixmap(scaled)

    # Preserve composite scaling on parent resize (matches original behavior)
    def resizeEvent(self, event):
        super().resizeEvent(event)
        if hasattr(self, "_full_canvas") and self._full_canvas is not None:
            scaled = self._full_canvas.scaled(
                self.label_displayImages.size(),
                Qt.KeepAspectRatio,
                Qt.SmoothTransformation
            )
            self.label_displayImages.setPixmap(scaled)

    # Optional reject hook (wired if buttonBox_close exists)
    def reject(self):
        # No special behavior beyond original; provided for compatibility with wiring
        pass
