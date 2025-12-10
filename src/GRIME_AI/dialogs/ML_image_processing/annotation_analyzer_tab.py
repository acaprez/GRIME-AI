import os
import json
from pathlib import Path

from openpyxl import Workbook

from PyQt5 import QtWidgets, uic, QtCore
from PyQt5.QtWidgets import QFileDialog, QListWidgetItem, QMessageBox

from GRIME_AI.GRIME_AI_Save_Utils import GRIME_AI_Save_Utils


# ======================================================================================================================
# ======================================================================================================================
# ===   ===   ===   ===   ===   ===   ===      class AnnotationAnalyzerTab       ===   ===   ===   ===   ===   ===   ===
# ======================================================================================================================
# ======================================================================================================================
class AnnotationAnalyzerTab(QtWidgets.QWidget):
    analyze_signal = QtCore.pyqtSignal(list)

    # ------------------------------------------------------------------------------------------------------------------
    # ------------------------------------------------------------------------------------------------------------------
    def __init__(self, parent=None):
        super().__init__(parent)
        ui_path = Path(__file__).parent / "annotation_analyzer_tab.ui"
        uic.loadUi(str(ui_path), self)

        self.transferred_items = set()
        self.original_jsons = []

        # Wire up signals
        self.pushButton_browse_annotation_root_folder.clicked.connect(self.browse_root_folder)
        self.pushButton_moveRight.clicked.connect(self.move_to_selected)
        self.pushButton_moveLeft.clicked.connect(self.move_to_available)
        self.pushButton_reset.clicked.connect(self.reset_lists)
        self.pushButton_reset.setStyleSheet("""
        QPushButton {
            background-color: darkred;
            color: white;
            font: bold 12pt;
        }
        QPushButton:disabled {
            background-color: lightgray;
            color: darkgray;
        }
        """)

        self.pushButton_analyze.clicked.connect(self.analyze)
        self.pushButton_analyze.setStyleSheet("""
        QPushButton {
            background-color: steelblue;
            color: white;
            font: bold 12pt;
        }
        QPushButton:disabled {
            background-color: lightgray;
            color: darkgray;
        }
        """)

        # Monitor changes in the selected listbox
        self.listWidget_selectedJsons.itemSelectionChanged.connect(self.updateAnalyzeButtonState)

        # Initialize button state
        self.updateAnalyzeButtonState()

    # ------------------------------------------------------------------------------------------------------------------
    # ------------------------------------------------------------------------------------------------------------------
    def updateAnalyzeButtonState(self):
        """Enable Analyze button only if there are items in the selected listbox."""
        has_items = self.listWidget_selectedJsons.count() > 0
        self.pushButton_analyze.setEnabled(has_items)

    # ------------------------------------------------------------------------------------------------------------------
    # ------------------------------------------------------------------------------------------------------------------
    def browse_root_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "Select Root Folder", os.getcwd())
        if folder:
            self.lineEdit_annotation_root_path.setText(folder)
            self.populate_available_jsons(Path(folder))

    # ------------------------------------------------------------------------------------------------------------------
    # ------------------------------------------------------------------------------------------------------------------
    def populate_available_jsons(self, root: Path):
        self.listWidget_availableJsons.clear()
        self.listWidget_selectedJsons.clear()
        self.transferred_items.clear()
        self.original_jsons.clear()

        for dirpath, _, filenames in os.walk(root):
            if "instances_default.json" in filenames:
                full_path = os.path.normpath(os.path.join(dirpath, "instances_default.json"))
                self.original_jsons.append(full_path)

        for path in sorted(self.original_jsons):
            self.listWidget_availableJsons.addItem(QListWidgetItem(path))

        self.updateAnalyzeButtonState()

    # ------------------------------------------------------------------------------------------------------------------
    # ------------------------------------------------------------------------------------------------------------------
    def move_to_selected(self):
        for item in self.listWidget_availableJsons.selectedItems():
            text = item.text()
            row = self.listWidget_availableJsons.row(item)
            self.listWidget_availableJsons.takeItem(row)
            if text not in self.transferred_items:
                self.listWidget_selectedJsons.addItem(text)
                self.transferred_items.add(text)
        self.updateAnalyzeButtonState()

    # ------------------------------------------------------------------------------------------------------------------
    # ------------------------------------------------------------------------------------------------------------------
    def move_to_available(self):
        for item in self.listWidget_selectedJsons.selectedItems():
            text = item.text()
            row = self.listWidget_selectedJsons.row(item)
            self.listWidget_selectedJsons.takeItem(row)
            if text in self.transferred_items:
                self.transferred_items.remove(text)
            self.listWidget_availableJsons.addItem(text)
        self.updateAnalyzeButtonState()

    # ------------------------------------------------------------------------------------------------------------------
    # ------------------------------------------------------------------------------------------------------------------
    def reset_lists(self):
        self.listWidget_availableJsons.clear()
        self.listWidget_selectedJsons.clear()
        self.transferred_items.clear()
        for path in sorted(self.original_jsons):
            self.listWidget_availableJsons.addItem(QListWidgetItem(path))
        self.updateAnalyzeButtonState()

    # ------------------------------------------------------------------------------------------------------------------
    # ------------------------------------------------------------------------------------------------------------------
    def analyze(self):
        selected = [self.listWidget_selectedJsons.item(i).text()
                    for i in range(self.listWidget_selectedJsons.count())]

        if not selected:
            QMessageBox.warning(self, "No Selection", "Please select at least one JSON file.")
            return

        category_sets = []
        for path in selected:
            try:
                with open(path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                categories = {cat["id"]: cat["name"] for cat in data.get("categories", [])}
                category_sets.append(categories)
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to read {path}:\n{e}")
                return

        # Deduplicate sets (optional, keep full list for export)
        unique_sets = []
        seen = set()
        for cats in category_sets:
            items = tuple(sorted(cats.items()))
            if items not in seen:
                seen.add(items)
                unique_sets.append(cats)

        # Display dialog
        display_text = "Unique category sets found:\n\n"
        for idx, cats in enumerate(unique_sets, start=1):
            display_text += f"Set {idx}:\n"
            for cid, name in cats.items():
                display_text += f"  ID {cid}: {name}\n"
            display_text += "\n"

        QMessageBox.information(self, "Annotation Analysis", display_text)

        # Export all selected datasets to XLSX
        self.export_to_xlsx(selected, category_sets)

    # ------------------------------------------------------------------------------------------------------------------
    # ------------------------------------------------------------------------------------------------------------------
    from openpyxl import Workbook
    import os

    # ------------------------------------------------------------------------------------------------------------------
    # ------------------------------------------------------------------------------------------------------------------
    def export_to_xlsx(self, selected_paths, category_sets):
        """
        Generate an XLSX file with dataset folder names in the first column,
        followed by id, label pairs for each category.
        Output goes into analysis_outputs folder inside the user's GRIME AI folder.
        """
        # Get user's GRIME AI folder
        grime_ai_folder = GRIME_AI_Save_Utils().get_users_GRIME_AI_folder()

        # Create analysis_outputs subfolder if it doesn't exist
        output_dir = os.path.join(grime_ai_folder, "analysis_outputs")
        os.makedirs(output_dir, exist_ok=True)

        # Prepare workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Annotation Categories"

        # Header row
        ws.cell(row=1, column=1, value="Dataset Folder")
        ws.cell(row=1, column=2, value="Categories (id, label, ...)")

        # Fill rows
        for row_idx, (path, cats) in enumerate(zip(selected_paths, category_sets), start=2):
            folder_name = Path(path).parent.name  # last folder containing the JSON
            ws.cell(row=row_idx, column=1, value=folder_name)
            col_idx = 2
            for cid, name in cats.items():
                ws.cell(row=row_idx, column=col_idx, value=cid)
                ws.cell(row=row_idx, column=col_idx + 1, value=name)
                col_idx += 2

        # Save file into analysis_outputs folder
        out_path = os.path.join(output_dir, "annotation_analysis.xlsx")
        wb.save(out_path)

        QMessageBox.information(self, "Export Complete",
                                f"XLSX file generated:\n{out_path}")
