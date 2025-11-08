#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# Author: John Edward Stranzl, Jr.
# Affiliation(s): University of Nebraska-Lincoln, Blade Vision Systems, LLC
# Contact: jstranzl2@huskers.unl.edu, johnstranzl@gmail.com
# Created: Mar 6, 2022
# License: Apache License, Version 2.0, http://www.apache.org/licenses/LICENSE-2.0

# green_image_generator.py

import os
from datetime import datetime, timedelta

import colorsys
import numpy as np
import cv2
from PIL import Image, ImageDraw

from openpyxl import Workbook
from openpyxl.styles import PatternFill


class GreenImageGenerator:
    """
    Generate solid-green, splotch-on-clay, and mask images
    with options for flat or textured backgrounds and spots.
    """

    def __init__(
        self,
        out_dir="Test_Images",
        width=1024,
        height=768,
        num_solids=64,
        num_splotch=32,
        splotch_count=50,
        splotch_rad=(20, 100),
        noise_amplitude=20,
        clay_bg=(160, 82, 45),
        hue_green=100,
    ):
        self.OUT_DIR = out_dir
        self.W = width
        self.H = height
        self.NUM_SOLIDS = num_solids
        self.NUM_SPLOTCH = num_splotch
        self.SPLOTCH_COUNT = splotch_count
        self.SPLOTCH_RAD = splotch_rad
        self.NOISE_AMPLITUDE = noise_amplitude
        self.CLAY_BG = clay_bg
        self.HUE_GREEN = hue_green

        self.solid_dir = None
        self.splotch_dir = None
        self.mask_dir = None

    @staticmethod
    def hsv_to_rgb8(h, s, v):
        r, g, b = colorsys.hsv_to_rgb(h / 360.0, s, v)
        return int(r * 255), int(g * 255), int(b * 255)

    @staticmethod
    def timestamp_suffix(base_time, index):
        ts = base_time + timedelta(minutes=index)
        return f"_{ts.strftime('%Y_%m_%d_%H%M%S')}".lower()

    def make_dirs(self):
        self.solid_dir = os.path.join(self.OUT_DIR, "solids")
        self.splotch_dir = os.path.join(self.OUT_DIR, "splotches")
        self.mask_dir = os.path.join(self.OUT_DIR, "masks")
        os.makedirs(self.solid_dir, exist_ok=True)
        os.makedirs(self.splotch_dir, exist_ok=True)
        os.makedirs(self.mask_dir, exist_ok=True)

    def generate_textured_image(self, base_color, size):
        arr = np.full((size[1], size[0], 3), base_color, dtype=np.float32)
        noise = (np.random.rand(size[1], size[0]) * 2 * self.NOISE_AMPLITUDE) - self.NOISE_AMPLITUDE
        noise = cv2.GaussianBlur(noise, (3, 3), 0)
        for c in range(3):
            arr[:, :, c] += noise
        np.clip(arr, 0, 255, out=arr)
        return Image.fromarray(arr.astype(np.uint8))

    def generate_textured_clay(self):
        base = np.full((self.H, self.W, 3), self.CLAY_BG, dtype=np.float32)
        fine_noise = (np.random.rand(self.H, self.W) * 40) - 20
        fine_noise = cv2.GaussianBlur(fine_noise, (3, 3), 0)
        med_noise = cv2.resize(
            np.random.randn(self.H // 32, self.W // 32) * 5,
            (self.W, self.H),
            interpolation=cv2.INTER_LINEAR,
        )
        med_noise = cv2.GaussianBlur(med_noise, (31, 31), 0)
        combined = fine_noise + med_noise
        for c in range(3):
            base[:, :, c] += combined
        np.clip(base, 0, 255, out=base)
        return Image.fromarray(base.astype(np.uint8))

    def generate_solid_images(self, texture="flat"):
        sats = np.linspace(0.2, 1.0, 8)
        vals = np.linspace(0.2, 1.0, 8)
        combos = [(s, v) for v in vals for s in sats]
        combos.insert(0, ("custom", "custom"))

        if texture == "flat":
            base_time = datetime.now().replace(hour=12, minute=0, second=0, microsecond=0)
        else:
            base_time = datetime.now().replace(hour=15, minute=0, second=0, microsecond=0)

        time_prefix = base_time.strftime("%Y%m%d_%H%M")
        csv_name = "Solid_Flat_Images_Info.csv" if texture == "flat" else "Solid_Textured_Images_Info.csv"
        xlsx_name = f"{time_prefix}_{csv_name.replace('.csv', '.xlsx')}"
        img_dir = os.path.join(self.solid_dir, texture)
        os.makedirs(img_dir, exist_ok=True)
        xlsx_path = os.path.join(img_dir, xlsx_name)

        header = [
            "filename", "attribute", "texture_type", "color_swatch", "base_RGB", "base_HSV",
            "mean_RGB", "mean_HSV", "std_RGB", "std_HSV",
            "GCC", "ExG", "VARI", "GLI"
        ]

        wb = Workbook()
        ws = wb.active
        ws.title = "Solid Image Data"
        ws.append(header)

        for idx, (s, v) in enumerate(combos):
            if s == "custom":
                base_rgb = (0, 255, 0)
                base_hsv_str = "custom"
            else:
                base_rgb = self.hsv_to_rgb8(self.HUE_GREEN, s, v)
                base_hsv_str = f"({self.HUE_GREEN}, {round(s, 3)}, {round(v, 3)})"

            fname = f"grimeaitest{self.timestamp_suffix(base_time, idx)}".lower()

            if texture == "flat":
                img = Image.new("RGB", (self.W, self.H), base_rgb)
                mean_rgb = base_rgb
                std_rgb = (0.0, 0.0, 0.0)
                mean_hsv = base_hsv_str if base_hsv_str == "custom" else (self.HUE_GREEN, round(s, 3), round(v, 3))
                std_hsv = (0.0, 0.0, 0.0)
            else:
                img = self.generate_textured_image(base_rgb, (self.W, self.H))
                arr = np.array(img).astype(np.float32)
                mean_rgb = tuple(np.round(np.mean(arr, axis=(0, 1)), 2))
                std_rgb = tuple(np.round(np.std(arr, axis=(0, 1)), 2))

                mean_hsv_raw = colorsys.rgb_to_hsv(*(c / 255 for c in mean_rgb))
                mean_hsv = (round(mean_hsv_raw[0] * 360, 2), round(mean_hsv_raw[1], 3), round(mean_hsv_raw[2], 3))

                std_hsv_raw = colorsys.rgb_to_hsv(*(c / 255 for c in std_rgb))
                std_hsv = (round(std_hsv_raw[0] * 360, 2), round(std_hsv_raw[1], 3), round(std_hsv_raw[2], 3))

            save_path = os.path.join(img_dir, fname + ".jpg")
            img.save(save_path, quality=100)

            r, g, b = mean_rgb
            rgb_sum = r + g + b
            gcc = round(g / rgb_sum, 4) if rgb_sum else 0.0
            exg = round(2 * g - r - b, 4)
            vari = round((g - r) / (g + r - b), 4) if (g + r - b) else 0.0
            gli = round((2 * g - r - b) / (2 * g + r + b), 4) if (2 * g + r + b) else 0.0

            row = [
                fname + ".jpg", "solid", texture, "", str(base_rgb), base_hsv_str,
                str(mean_rgb), str(mean_hsv), str(std_rgb), str(std_hsv),
                gcc, exg, vari, gli
            ]
            ws.append(row)

            swatch_col = header.index("color_swatch") + 1
            hex_color = "%02X%02X%02X" % base_rgb
            ws.cell(row=ws.max_row, column=swatch_col).fill = PatternFill(
                start_color=hex_color, end_color=hex_color, fill_type="solid"
            )

        wb.save(xlsx_path)
        print(f"Solid XLSX created → {xlsx_path}")

    def generate_splotch_images(self, background_texture="flat", splotch_texture="flat"):
        if background_texture == "flat":
            base_time = datetime.now().replace(hour=18, minute=0, second=0, microsecond=0)
        else:
            base_time = datetime.now().replace(hour=21, minute=0, second=0, microsecond=0)

        subfolder = f"{background_texture}_{splotch_texture}"
        img_dir = os.path.join(self.splotch_dir, subfolder)
        os.makedirs(img_dir, exist_ok=True)
        xlsx_path = os.path.join(img_dir, f"Splotch_{subfolder}_Images.xlsx")

        wb = Workbook()
        ws = wb.active
        ws.title = "Splotch Image Data"

        header = ["filename", "background", "splotch", "Base_Swatch", "Base"]
        if background_texture == "textured":
            header += ["Background Mean RGB", "Background Mean HSV", "Background Std RGB", "Background Std HSV"]
        for i in range(1, 4):
            header += [f"Splotch {i} Swatch", f"Splotch {i} RGB", f"Splotch {i} HSV"]
            if splotch_texture == "textured":
                header += [
                    f"Splotch {i} Mean RGB", f"Splotch {i} Mean HSV",
                    f"Splotch {i} Std RGB", f"Splotch {i} Std HSV"
                ]
        ws.append(header)

        for idx in range(self.NUM_SPLOTCH):
            # background
            bg_color = self.CLAY_BG
            if background_texture == "flat":
                bg = Image.new("RGB", (self.W, self.H), bg_color)
            else:
                bg = self.generate_textured_clay()
                arr_bg = np.array(bg).astype(np.float32)
                mean_rgb_bg = tuple(np.round(np.mean(arr_bg, axis=(0, 1)), 2))
                std_rgb_bg = tuple(np.round(np.std(arr_bg, axis=(0, 1)), 2))
                mean_hsv_raw = colorsys.rgb_to_hsv(*(c / 255 for c in mean_rgb_bg))
                mean_hsv_bg = (
                    round(mean_hsv_raw[0] * 360, 2),
                    round(mean_hsv_raw[1], 3),
                    round(mean_hsv_raw[2], 3),
                )
                std_hsv_raw = colorsys.rgb_to_hsv(*(c / 255 for c in std_rgb_bg))
                std_hsv_bg = (
                    round(std_hsv_raw[0] * 360, 2),
                    round(std_hsv_raw[1], 3),
                    round(std_hsv_raw[2], 3),
                )

            hsvs = [
                (self.HUE_GREEN, np.random.uniform(0.2, 1), np.random.uniform(0.2, 1))
                for _ in range(3)
            ]
            rgbs = [self.hsv_to_rgb8(*hsv) for hsv in hsvs]

            patch_stats = []
            for rgb, hsv in zip(rgbs, hsvs):
                r = np.random.randint(*self.SPLOTCH_RAD)
                x = np.random.randint(r, self.W - r)
                y = np.random.randint(r, self.H - r)
                patch = (
                    Image.new("RGB", (2 * r, 2 * r), rgb)
                    if splotch_texture == "flat"
                    else self.generate_textured_image(rgb, (2 * r, 2 * r))
                )

                if splotch_texture == "textured":
                    arr = np.array(patch).astype(np.float32)
                    mean_rgb = tuple(np.round(np.mean(arr, axis=(0, 1)), 2))
                    std_rgb = tuple(np.round(np.std(arr, axis=(0, 1)), 2))
                    mean_hsv_raw = colorsys.rgb_to_hsv(*(c / 255 for c in mean_rgb))
                    std_hsv_raw = colorsys.rgb_to_hsv(*(c / 255 for c in std_rgb))
                    mean_hsv = (
                        round(mean_hsv_raw[0] * 360, 2),
                        round(mean_hsv_raw[1], 3),
                        round(mean_hsv_raw[2], 3),
                    )
                    std_hsv = (
                        round(std_hsv_raw[0] * 360, 2),
                        round(std_hsv_raw[1], 3),
                        round(std_hsv_raw[2], 3),
                    )
                    patch_stats.append((rgb, hsv, mean_rgb, mean_hsv, std_rgb, std_hsv))
                else:
                    patch_stats.append((rgb, hsv, None, None, None, None))

                mask = Image.new("L", (2 * r, 2 * r), 0)
                ImageDraw.Draw(mask).ellipse((0, 0, 2 * r, 2 * r), fill=255)
                bg.paste(patch, (x - r, y - r), mask)

            fname = f"grimeaitest{self.timestamp_suffix(base_time, idx)}.jpg".lower()
            bg.save(os.path.join(img_dir, fname), quality=100)

            row = [fname, background_texture, splotch_texture, "", str(bg_color)]
            if background_texture == "textured":
                row += [str(mean_rgb_bg), str(mean_hsv_bg), str(std_rgb_bg), str(std_hsv_bg)]
            for rgb, hsv, mean_rgb, mean_hsv, std_rgb, std_hsv in patch_stats:
                hsv_str = f"({self.HUE_GREEN}, {round(hsv[1], 3)}, {round(hsv[2], 3)})"
                row += ["", str(rgb), hsv_str]
                if splotch_texture == "textured":
                    row += [str(mean_rgb), str(mean_hsv), str(std_rgb), str(std_hsv)]
            ws.append(row)

            # fill swatches
            base_col = header.index("Base_Swatch") + 1
            hex_bg = "%02X%02X%02X" % bg_color
            ws.cell(row=ws.max_row, column=base_col).fill = PatternFill(
                start_color=hex_bg, end_color=hex_bg, fill_type="solid"
            )
            for i, (rgb, *_rest) in enumerate(patch_stats):
                label = f"Splotch {i+1} Swatch"
                col = header.index(label) + 1
                hex_c = "%02X%02X%02X" % rgb
                ws.cell(row=ws.max_row, column=col).fill = PatternFill(
                    start_color=hex_c, end_color=hex_c, fill_type="solid"
                )

        wb.save(xlsx_path)
        print(f"Splotch XLSX created → {xlsx_path}")

    def generate_mask_images(self):
        os.makedirs(self.mask_dir, exist_ok=True)
        base_time = datetime.now().replace(hour=12, minute=0, second=0, microsecond=0)

        masks = [
            ("mask_3c_white", (255, 255, 255), "RGB"),
            ("mask_3c_black", (0, 0, 0), "RGB"),
            ("mask_1c_white", 255, "L"),
            ("mask_1c_black", 0, "L"),
        ]

        for idx, (basename, fill, mode) in enumerate(masks):
            img = Image.new(mode, (self.W, self.H), fill)
            fname = (basename + self.timestamp_suffix(base_time, idx)).lower()
            for ext in [".png", ".jpg", ".tif"]:
                img.save(os.path.join(self.mask_dir, fname + ext), quality=100)

        print(f"Mask images saved to {self.mask_dir}")

    def generate_all(self):
        """
        Convenience method to generate:
        - Solid images (both flat and textured)
        - Splotch images (all texture permutations)
        - Mask images
        """
        self.make_dirs()
        self.generate_solid_images(texture="flat")
        self.generate_solid_images(texture="textured")

        for bg in ["flat", "textured"]:
            for sp in ["flat", "textured"]:
                self.generate_splotch_images(background_texture=bg, splotch_texture=sp)

        self.generate_mask_images()
